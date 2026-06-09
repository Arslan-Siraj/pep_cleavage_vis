#!/usr/bin/env python3
"""
run_analysis_lollipop.py

Presentation-friendly lollipop visualization for Arabidopsis AT3G09260.1
(PYK10 / BGLU23 / LEB / PSR3.1) cleavage-site evidence.

This script reads the raw Excel files, extracts PYK10 rows from three Arabidopsis
datasets, converts them to a common protein-coordinate system, clusters nearby
cleavage sites, and generates a clean lollipop plot.

Main input files expected in --input-dir:
  1. Arabidopsis semitryptome Excel file
  2. HUNTER_data_shoot_MJ Excel file
  3. ProteinBased_www_P10'P10 Excel file

The script is robust to copied/renamed files such as:
  Arabidopsis_semitryptome_23112020.xlsx
  Arabidopsis_semitryptome_23112020(2).xlsx
  HUNTER_data_shoot_MJ.xlsx
  HUNTER_data_shoot_MJ(1).xlsx
  ProteinBased_www_P10'P10 1.xlsx
  ProteinBased_www_P10'P10 1(1).xlsx
  ProteinBased_www_P10d_P10_1.xlsx

Usage:
  python scripts/run_analysis_lollipop.py --input-dir data --output-dir output

Outputs:
  output/tables/pyk10_semitryptome.csv
  output/tables/pyk10_hunter_mj.csv
  output/tables/pyk10_metacaspase.csv
  output/tables/pyk10_lollipop_clusters.csv
  output/tables/pyk10_lollipop_points.csv
  output/plots/pyk10_lollipop_multisource_hotspots.png
  output/plots/pyk10_lollipop_multisource_hotspots.pdf
  output/README_lollipop.md
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------

TARGET_PROTEIN = "AT3G09260.1"
TARGET_UNIPROT = "Q9SR37"
TARGET_GENE_TERMS = ["PYK10", "BGLU23", "LEB", "PSR3.1"]
DEFAULT_LENGTH = 525

LOGFC_COLS = [
    "LogFC\n5 vs 0",
    "LogFC \n30 vs 0",
    "LogFC \n60 vs 0",
    "LogFC\n180 vs 0",
    "LogFC\n480 vs 0",
]

# These are preferred names, but the resolver below also searches flexibly.
INPUT_FILENAMES = {
    "semitryptome": "Arabidopsis_semitryptome_23112020.xlsx",
    "hunter_arabidopsis": "HUNTER_data_shoot_MJ.xlsx",
    "metacaspase": "ProteinBased_www_P10d_P10_1.xlsx",
}


# ---------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------

def safe_numeric(x) -> pd.Series:
    return pd.to_numeric(x, errors="coerce")


def clean_peptide_sequence(seq: object) -> str:
    """Remove modification notation and non-amino-acid characters."""
    seq = re.sub(r"\[[^\]]+\]", "", str(seq))
    return re.sub(r"[^A-Z]", "", seq.upper())


def peptide_span_from_start(peptide: object, start: object) -> Tuple[Optional[int], Optional[int]]:
    """Calculate peptide start/end coordinates from start coordinate and sequence."""
    try:
        st = int(float(start))
    except Exception:
        return None, None

    seq = clean_peptide_sequence(peptide)
    if not seq:
        return st, st

    return st, st + len(seq) - 1


def normalize_filename(name: str) -> str:
    """
    Normalize filename for robust matching.
    Example:
      ProteinBased_www_P10'P10 1(1).xlsx
      ProteinBased_www_P10d_P10_1.xlsx
    both normalize to similar alphanumeric strings.
    """
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def resolve_input_files(input_dir: Path) -> Dict[str, Path]:
    """
    Resolve input Excel files robustly.

    This avoids failures caused by apostrophes, spaces, copied suffixes such as (1),
    and sanitized filenames such as P10d_P10.
    """
    input_dir = Path(input_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")

    xlsx_files = sorted(input_dir.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx files found in input directory: {input_dir}")

    patterns = {
        "semitryptome": [
            "arabidopsissemitryptome",
            "semitryptome",
        ],
        "hunter_arabidopsis": [
            "hunterdatashootmj",
            "huntershootmj",
            "shootmj",
        ],
        "metacaspase": [
            "proteinbasedwwwp10p101",
            "proteinbasedp10p10",
            "proteinbasedwwwp10dp101",
            "proteinbased",
            "p10p10",
            "p10dp10",
        ],
    }

    resolved: Dict[str, Path] = {}

    for key, preferred in INPUT_FILENAMES.items():
        preferred_path = input_dir / preferred
        if preferred_path.exists():
            resolved[key] = preferred_path
            continue

        candidates = []
        for f in xlsx_files:
            nf = normalize_filename(f.name)
            if any(pat in nf for pat in patterns[key]):
                candidates.append(f)

        if not candidates:
            available = "\n".join(f"  - {f.name}" for f in xlsx_files)
            raise FileNotFoundError(
                f"Could not find required Excel file for '{key}'.\n"
                f"Input directory: {input_dir}\n"
                f"Preferred filename: {preferred}\n\n"
                f"Available .xlsx files:\n{available}\n\n"
                f"Fix: rename the file, or edit INPUT_FILENAMES in this script."
            )

        # Prefer the shortest matching filename, then alphabetic order.
        resolved[key] = sorted(candidates, key=lambda p: (len(p.name), p.name.lower()))[0]

    print("Resolved input files:")
    for key, path in resolved.items():
        print(f"  {key}: {path.name}")

    return resolved


def ensure_output_dirs(output_dir: Path) -> Tuple[Path, Path]:
    output_dir = Path(output_dir)
    plots = output_dir / "plots"
    tables = output_dir / "tables"
    plots.mkdir(parents=True, exist_ok=True)
    tables.mkdir(parents=True, exist_ok=True)
    return plots, tables


def unique_headers(raw_headers: Sequence[object]) -> List[str]:
    """
    Make duplicate Excel headers unique, similar to pandas behavior.
    """
    seen = {}
    headers = []
    for h in raw_headers:
        base = str(h) if h is not None else "Unnamed"
        if base in seen:
            seen[base] += 1
            headers.append(f"{base}.{seen[base]}")
        else:
            seen[base] = 0
            headers.append(base)
    return headers


# ---------------------------------------------------------------------
# Excel readers
# ---------------------------------------------------------------------

def read_semitryptome(path: Path) -> pd.DataFrame:
    """
    Read the Arabidopsis semitryptome file and retain only PYK10/BGLU23 rows.

    Expected relevant columns:
      Protein, Gene, Description, CleavageSite, P10, P10', PSM, #Samples, Peps
    """
    wanted = [
        "Protein", "Gene", "Description", "Representative", "CleavageSite",
        "P10", "P10'", "PSM", "PSM_unique", "#Peps", "Peps", "#Samples",
        "TargetP_pred", "TargetP_CS", "TargetP_prob"
    ]

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = unique_headers(raw_headers)
    idx = {h: i for i, h in enumerate(headers)}

    missing = [c for c in ["Protein", "Gene", "Description", "CleavageSite", "PSM", "Peps"] if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Missing semitryptome columns in {path.name}: {missing}")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        protein = r[idx["Protein"]]
        gene = r[idx["Gene"]]
        desc = r[idx["Description"]]

        is_target = (
            str(protein).upper() == TARGET_UNIPROT
            or str(gene).upper() == "BGLU23"
            or re.search(r"beta-glucosidase 23", str(desc), flags=re.I) is not None
        )

        if is_target:
            rows.append({c: r[idx[c]] if c in idx else None for c in wanted})

    wb.close()

    out = pd.DataFrame(rows)
    if out.empty:
        print("WARNING: no PYK10/BGLU23 rows found in semitryptome.")
        return out

    out["dataset"] = "Semitryptome"
    out["cleavage_site"] = safe_numeric(out["CleavageSite"]).astype("Int64")
    out["psm"] = safe_numeric(out["PSM"])
    out["samples"] = safe_numeric(out["#Samples"])
    out["peptide"] = out["Peps"].astype(str).str.split(";").str[0].map(clean_peptide_sequence)
    out["start"], out["end"] = zip(*[
        peptide_span_from_start(p, s) for p, s in zip(out["peptide"], out["cleavage_site"])
    ])
    out["weight"] = out["psm"].fillna(1)

    return out


def read_hunter_arabidopsis(path: Path, sheet_name: str = "Only significant peptides") -> pd.DataFrame:
    """
    Read Arabidopsis HUNTER methyl-jasmonate data and retain PYK10 rows.

    If the requested sheet is absent, the script tries 'All peptides'.
    """
    wb = load_workbook(path, read_only=True, data_only=True)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "All peptides" in wb.sheetnames:
        print(f"WARNING: sheet '{sheet_name}' not found in {path.name}; using 'All peptides'.")
        ws = wb["All peptides"]
    else:
        wb.close()
        raise ValueError(f"Could not find HUNTER sheet in {path.name}. Available sheets: {wb.sheetnames}")

    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = unique_headers(raw_headers)
    idx = {h: i for i, h in enumerate(headers)}

    required = ["protein", "gene", "peptide"]
    missing = [c for c in required if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Missing HUNTER columns in {path.name}: {missing}")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        protein = r[idx["protein"]]
        gene = r[idx["gene"]]

        is_target = (
            str(protein).upper() == TARGET_PROTEIN.upper()
            or re.search(r"BGLU23|PYK10|LEB|PSR3", str(gene), flags=re.I) is not None
        )

        if is_target:
            rows.append({h: r[i] if i < len(r) else None for i, h in enumerate(headers)})

    wb.close()

    out = pd.DataFrame(rows)
    if out.empty:
        print("WARNING: no PYK10/BGLU23 rows found in HUNTER data.")
        return out

    out["dataset"] = "HUNTER MJ"

    # In some HUNTER tables, "cleavage_site" contains a sequence string and
    # numeric coordinates are stored in "first_aa_position".
    cs = safe_numeric(out["cleavage_site"]) if "cleavage_site" in out.columns else pd.Series(np.nan, index=out.index)
    if cs.isna().all() and "first_aa_position" in out.columns:
        cs = safe_numeric(out["first_aa_position"])

    out["cleavage_site"] = cs.astype("Int64")
    out["peptide"] = out["peptide"].astype(str).map(clean_peptide_sequence)
    out["start"], out["end"] = zip(*[
        peptide_span_from_start(p, s) for p, s in zip(out["peptide"], out["cleavage_site"])
    ])

    for c in LOGFC_COLS:
        if c in out.columns:
            out[c] = safe_numeric(out[c])

    logfc_present = [c for c in LOGFC_COLS if c in out.columns]
    if logfc_present:
        out["max_abs_logfc"] = out[logfc_present].abs().max(axis=1, skipna=True)
        out["mean_logfc"] = out[logfc_present].mean(axis=1, skipna=True)
    else:
        out["max_abs_logfc"] = np.nan
        out["mean_logfc"] = np.nan

    out["weight"] = out["max_abs_logfc"].fillna(0)

    return out


def read_metacaspase(path: Path) -> pd.DataFrame:
    """
    Read the ProteinBased_www_P10'P10 metacaspase matrix and retain only rows
    whose PRIMARY Protein column is AT3G09260.1.

    Important decision:
      Some rows have Protein = another BGLU homolog, for example AT3G21370.1,
      but the Isoforms column also contains AT3G09260.1. These are shared or
      homologous peptide mappings, not primary PYK10 rows. For the requested
      PYK10-specific plot, they are excluded.

    Strict rule used here:
      keep a row only if Protein == AT3G09260.1
    """
    wb = load_workbook(path, read_only=True, data_only=True)

    if "Data" in wb.sheetnames:
        ws = wb["Data"]
    else:
        ws = wb[wb.sheetnames[0]]
        print(f"WARNING: sheet 'Data' not found in {path.name}; using '{ws.title}'.")

    header_row = 4
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    headers = unique_headers(raw_headers)
    idx = {h: i for i, h in enumerate(headers)}

    required = ["Protein", "Description", "Start", "End", "Sequence"]
    missing = [c for c in required if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Missing metacaspase columns in {path.name}: {missing}")

    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        prot = str(r[idx["Protein"]]).strip().upper() if r[idx["Protein"]] is not None else ""

        # Primary protein identifier only. Do not use Description or Isoforms for
        # final PYK10-specific extraction, because they can include shared/homologous mappings.
        if prot == TARGET_PROTEIN.upper():
            rows.append({h: r[i] if i < len(r) else None for i, h in enumerate(headers)})

    wb.close()

    out = pd.DataFrame(rows)
    if out.empty:
        print("WARNING: no primary AT3G09260.1 rows found in metacaspase matrix.")
        return out

    out["dataset"] = "Metacaspase matrix"
    out["cleavage_site"] = safe_numeric(out["Start"]).astype("Int64")
    out["peptide"] = out["Sequence"].astype(str).map(clean_peptide_sequence)
    out["start"] = safe_numeric(out["Start"])
    out["end"] = safe_numeric(out["End"])

    score_cols = [c for c in out.columns if str(c).startswith("Score")]
    if score_cols:
        for c in score_cols:
            out[c] = safe_numeric(out[c])
        out["metacaspase_score_max"] = out[score_cols].max(axis=1, skipna=True)
    else:
        out["metacaspase_score_max"] = np.nan

    out["weight"] = out["metacaspase_score_max"].fillna(0)

    print("  Metacaspase final extracted Protein values:")
    for protein_id in sorted(out["Protein"].astype(str).unique()):
        print(f"    - {protein_id}")

    return out


# ---------------------------------------------------------------------
# Lollipop plotting
# ---------------------------------------------------------------------

def build_lollipop_points(
    semi: pd.DataFrame,
    hunter: pd.DataFrame,
    meta: pd.DataFrame,
    aggregate_sites: bool = True,
) -> pd.DataFrame:
    """
    Convert the three extracted PYK10 tables into one plotting table.

    Weight meaning:
      Semitryptome: PSM
      HUNTER MJ: maximum absolute logFC
      Metacaspase matrix: maximum metacaspase score

    By default this function collapses exact duplicate coordinates, so the plot
    has one dot per dataset per cleavage coordinate. This avoids a misleading
    plot in which repeated peptide observations at the same coordinate look like
    many independent cleavage sites.
    """
    parts = []

    if not semi.empty:
        tmp = semi.copy()
        tmp["dataset"] = "Semitryptome"
        tmp["weight"] = safe_numeric(tmp.get("psm", tmp.get("PSM", 1))).fillna(1)
        parts.append(tmp[["cleavage_site", "dataset", "weight", "peptide", "start", "end"]])

    if not hunter.empty:
        tmp = hunter.copy()
        tmp["dataset"] = "HUNTER MJ"
        tmp["weight"] = safe_numeric(tmp.get("max_abs_logfc", 0)).fillna(0)
        parts.append(tmp[["cleavage_site", "dataset", "weight", "peptide", "start", "end"]])

    if not meta.empty:
        tmp = meta.copy()
        tmp["dataset"] = "Metacaspase matrix"
        tmp["weight"] = safe_numeric(tmp.get("metacaspase_score_max", 0)).fillna(0)
        parts.append(tmp[["cleavage_site", "dataset", "weight", "peptide", "start", "end"]])

    if not parts:
        return pd.DataFrame(columns=["cleavage_site", "dataset", "weight", "peptide", "start", "end"])

    combined = pd.concat(parts, ignore_index=True)
    combined["cleavage_site"] = safe_numeric(combined["cleavage_site"])
    combined = combined.dropna(subset=["cleavage_site"]).copy()
    combined["cleavage_site"] = combined["cleavage_site"].astype(int)
    combined["weight"] = safe_numeric(combined["weight"]).fillna(0)

    if not aggregate_sites:
        combined["n_observations"] = 1
        return combined

    def first_nonempty(x):
        vals = [str(v) for v in x if str(v) and str(v).lower() != "nan"]
        return ";".join(vals[:3]) if vals else ""

    aggregated = (
        combined
        .groupby(["dataset", "cleavage_site"], as_index=False)
        .agg(
            weight=("weight", "max"),
            n_observations=("weight", "size"),
            peptide=("peptide", first_nonempty),
            start=("start", "min"),
            end=("end", "max"),
        )
    )
    return aggregated


def build_cluster_summary(points: pd.DataFrame, gap: int = 10) -> pd.DataFrame:
    """
    Group nearby cleavage sites into processing hotspots.

    Rule:
      After sorting unique cleavage positions, a position joins the current cluster
      if it is within `gap` amino acids of the previous site.
    """
    if points.empty:
        return pd.DataFrame()

    sites = sorted(points["cleavage_site"].unique())

    clusters = []
    current = []
    for site in sites:
        if not current or site - current[-1] <= gap:
            current.append(site)
        else:
            clusters.append(current)
            current = [site]
    if current:
        clusters.append(current)

    rows = []
    for i, cluster_sites in enumerate(clusters, start=1):
        start = min(cluster_sites)
        end = max(cluster_sites)
        sub = points[points["cleavage_site"].between(start, end)]

        support = {}
        supporting_datasets = 0
        for dataset in ["Semitryptome", "HUNTER MJ", "Metacaspase matrix"]:
            n = sub[sub["dataset"] == dataset]["cleavage_site"].nunique()
            support[dataset] = int(n)
            if n > 0:
                supporting_datasets += 1

        rows.append({
            "cluster_id": f"C{i}",
            "cluster_start": start,
            "cluster_end": end,
            "cluster_center": float(np.mean(cluster_sites)),
            "sites": ";".join(map(str, cluster_sites)),
            "datasets_supporting": supporting_datasets,
            "Semitryptome": support["Semitryptome"],
            "HUNTER MJ": support["HUNTER MJ"],
            "Metacaspase matrix": support["Metacaspase matrix"],
            "n_total_sites": len(cluster_sites),
        })

    return pd.DataFrame(rows)


def infer_protein_length(semi: pd.DataFrame, hunter: pd.DataFrame, meta: pd.DataFrame) -> int:
    """
    Infer plotting length from peptide ends and HUNTER relative positions.
    """
    length_candidates = []

    if not hunter.empty and "preceding_aa_position_relative" in hunter.columns and "first_aa_position" in hunter.columns:
        rel = safe_numeric(hunter["preceding_aa_position_relative"])
        first = safe_numeric(hunter["first_aa_position"])
        valid = (rel > 0) & first.notna()
        length_candidates.extend((first[valid] / (rel[valid] / 100)).dropna().tolist())

    max_end = 0
    for df in [semi, hunter, meta]:
        if not df.empty and "end" in df.columns:
            m = safe_numeric(df["end"]).max()
            if pd.notna(m):
                max_end = max(max_end, int(m))

    if length_candidates:
        estimated = int(round(float(np.nanmedian(length_candidates))))
        return max(estimated, max_end, DEFAULT_LENGTH)

    return max(max_end, DEFAULT_LENGTH)


def plot_lollipop_multisource_hotspots(
    points: pd.DataFrame,
    cluster_summary: pd.DataFrame,
    protein_len: int,
    out_png: Path,
    out_pdf: Path,
    shade_min_dataset_support: int = 2,
    title: str = "AT3G09260.1 (PYK10 / BGLU23 / LEB / PSR3.1): cleavage-site evidence",
) -> None:
    """
    Generate a clean presentation-ready lollipop plot.

    Features:
      - one horizontal track per dataset
      - dot at each cleavage coordinate
      - marker size scaled by evidence strength
      - faint lollipop stems
      - background shading for clusters/hotspots supported by >= 2 datasets
    """
    if points.empty:
        raise ValueError("No points available for lollipop plot.")

    y_positions = {
        "Semitryptome": 3,
        "HUNTER MJ": 2,
        "Metacaspase matrix": 1,
    }

    fig, ax = plt.subplots(figsize=(15, 5.8))

    # Background cluster shading
    if not cluster_summary.empty:
        shaded = cluster_summary[cluster_summary["datasets_supporting"] >= shade_min_dataset_support]
        for _, row in shaded.iterrows():
            ax.axvspan(
                row["cluster_start"],
                row["cluster_end"],
                alpha=0.10,
                zorder=0,
            )
            ax.text(
                row["cluster_center"],
                3.47,
                row["cluster_id"],
                ha="center",
                va="bottom",
                fontsize=8,
            )

    # Lollipop tracks
    for dataset, y in y_positions.items():
        sub = points[points["dataset"] == dataset].copy()
        if sub.empty:
            continue

        max_weight = sub["weight"].max()
        if pd.isna(max_weight) or max_weight <= 0:
            max_weight = 1.0

        # Keep marker size readable but not too dominant.
        sub["marker_size"] = 30 + 85 * (sub["weight"] / max_weight)

        # Faint stems
        for x in sub["cleavage_site"]:
            ax.plot(
                [x, x],
                [y - 0.16, y + 0.16],
                linewidth=0.65,
                alpha=0.30,
                zorder=1,
            )

        ax.scatter(
            sub["cleavage_site"],
            np.full(len(sub), y),
            s=sub["marker_size"],
            alpha=0.75,
            label=dataset,
            zorder=2,
        )

    ax.set_xlim(0, protein_len + 10)
    ax.set_ylim(0.45, 3.75)
    ax.set_yticks([3, 2, 1])
    ax.set_yticklabels(["Semitryptome", "HUNTER MJ", "Metacaspase matrix"])
    ax.set_xlabel("AT3G09260.1 / PYK10 residue coordinate")
    ax.set_title(title)

    ax.grid(axis="x", alpha=0.23)
    ax.legend(
        loc="upper center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, 1.12),
    )

    fig.tight_layout()
    fig.savefig(out_png, dpi=300)
    fig.savefig(out_pdf)
    plt.close(fig)


def plot_cluster_matrix(cluster_summary: pd.DataFrame, out_png: Path, out_pdf: Path) -> None:
    """
    Optional companion plot: simple matrix showing which clusters are supported by which datasets.
    """
    if cluster_summary.empty:
        return

    keep = cluster_summary[
        (cluster_summary["datasets_supporting"] >= 2)
        | (cluster_summary["HUNTER MJ"] > 0)
    ].copy()

    if keep.empty:
        keep = cluster_summary.copy()

    dataset_cols = ["Semitryptome", "HUNTER MJ", "Metacaspase matrix"]
    keep["label"] = keep.apply(
        lambda r: f"{r['cluster_id']} ({int(r['cluster_start'])}-{int(r['cluster_end'])})",
        axis=1,
    )

    fig, ax = plt.subplots(figsize=(8.5, max(4.5, 0.35 * len(keep))))

    for i, (_, row) in enumerate(keep.iterrows()):
        for j, dataset in enumerate(dataset_cols):
            value = int(row[dataset])
            size = 25 + 70 * value if value > 0 else 8
            ax.scatter(j, i, s=size, alpha=0.75 if value else 0.20)
            if value:
                ax.text(j, i, str(value), ha="center", va="center", fontsize=8)

    ax.set_xticks(range(len(dataset_cols)))
    ax.set_xticklabels(dataset_cols)
    ax.set_yticks(range(len(keep)))
    ax.set_yticklabels(keep["label"])
    ax.invert_yaxis()
    ax.set_xlabel("Dataset")
    ax.set_ylabel("Cleavage-region cluster")
    ax.set_title("PYK10 cluster-level evidence matrix")
    ax.grid(alpha=0.2)

    fig.tight_layout()
    fig.savefig(out_png, dpi=300)
    fig.savefig(out_pdf)
    plt.close(fig)


def write_readme(
    output_dir: Path,
    sem: pd.DataFrame,
    hunter: pd.DataFrame,
    meta: pd.DataFrame,
    points: pd.DataFrame,
    cluster_summary: pd.DataFrame,
    cluster_gap: int,
) -> None:
    md = f"""# PYK10 lollipop cleavage-site visualization

This output was generated by `run_analysis_lollipop.py`.

## Objective

Visualize cleavage-site evidence for **AT3G09260.1 / PYK10 / BGLU23 / LEB / PSR3.1** as a clean lollipop plot, integrating three Arabidopsis datasets:

1. Arabidopsis semitryptome
2. HUNTER shoot methyl-jasmonate N-terminomics
3. Protein-based metacaspase substrate matrix

## Extracted rows

| Dataset | Rows extracted |
|---|---:|
| Semitryptome | {len(sem)} |
| HUNTER MJ | {len(hunter)} |
| Metacaspase matrix | {len(meta)} |
| Combined lollipop points | {len(points)} |
| Cleavage-site clusters | {len(cluster_summary)} |

## Plot interpretation

- **X-axis**: residue coordinate in AT3G09260.1 / PYK10.
- **Y-axis**: dataset source.
- **Dot position**: exact cleavage or peptide-start coordinate.
- **Dot size**: evidence strength.
  - Semitryptome: PSM support.
  - HUNTER MJ: maximum absolute log2 fold-change.
  - Metacaspase matrix: maximum substrate score.
- **Shaded bands**: cleavage-site clusters supported by at least two datasets.

Clusters were defined using a simple reproducible rule: adjacent cleavage sites are assigned to the same cluster when they are within **{cluster_gap} amino acids** of the previous site after sorting.

## Main files

- `plots/pyk10_lollipop_multisource_hotspots.png`
- `plots/pyk10_lollipop_multisource_hotspots.pdf`
- `plots/pyk10_cluster_evidence_matrix.png`
- `tables/pyk10_lollipop_points.csv`
- `tables/pyk10_lollipop_clusters.csv`

## Recommended presentation wording

Exact cleavage coordinates were retained, but nearby sites were additionally grouped into reproducible {cluster_gap}-aa clusters to highlight processing regions. This avoids overinterpreting adjacent N-terminal peptide starts as independent events while preserving the residue-level evidence in the output tables.
"""
    (output_dir / "README_lollipop.md").write_text(md, encoding="utf-8")


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a lollipop cleavage-site evidence plot for Arabidopsis PYK10."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("data"),
        help="Directory containing the input Excel files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output_lollipop"),
        help="Directory for output tables and plots.",
    )
    parser.add_argument(
        "--cluster-gap",
        type=int,
        default=10,
        help="Maximum gap in amino acids for grouping adjacent sites into clusters.",
    )
    parser.add_argument(
        "--shade-min-dataset-support",
        type=int,
        default=2,
        help="Minimum number of datasets supporting a cluster for background shading.",
    )
    parser.add_argument(
        "--plot-every-row",
        action="store_true",
        help=(
            "Plot every extracted row instead of aggregating exact duplicate cleavage "
            "positions. Default is one dot per dataset per cleavage coordinate."
        ),
    )
    args = parser.parse_args()

    plots_dir, tables_dir = ensure_output_dirs(args.output_dir)
    paths = resolve_input_files(args.input_dir)

    print("\nReading and extracting PYK10 rows...")
    sem = read_semitryptome(paths["semitryptome"])
    hunter = read_hunter_arabidopsis(paths["hunter_arabidopsis"])
    meta = read_metacaspase(paths["metacaspase"])

    print(f"  Semitryptome rows: {len(sem)}")
    print(f"  HUNTER MJ rows: {len(hunter)}")
    print(f"  Metacaspase matrix rows: {len(meta)}")

    def _n_unique_sites(df):
        if df.empty or "cleavage_site" not in df.columns:
            return 0
        return safe_numeric(df["cleavage_site"]).dropna().astype(int).nunique()

    print("  Unique exact cleavage coordinates:")
    print(f"    Semitryptome: {_n_unique_sites(sem)}")
    print(f"    HUNTER MJ: {_n_unique_sites(hunter)}")
    print(f"    Metacaspase matrix: {_n_unique_sites(meta)}")

    # Save extracted source tables.
    sem.to_csv(tables_dir / "pyk10_semitryptome.csv", index=False)
    hunter.to_csv(tables_dir / "pyk10_hunter_mj.csv", index=False)
    meta.to_csv(tables_dir / "pyk10_metacaspase.csv", index=False)

    points = build_lollipop_points(sem, hunter, meta, aggregate_sites=not args.plot_every_row)
    cluster_summary = build_cluster_summary(points, gap=args.cluster_gap)
    protein_len = infer_protein_length(sem, hunter, meta)

    points.to_csv(tables_dir / "pyk10_lollipop_points.csv", index=False)
    cluster_summary.to_csv(tables_dir / "pyk10_lollipop_clusters.csv", index=False)

    print(f"  Lollipop points: {len(points)}")
    print(f"  Clusters: {len(cluster_summary)}")
    print(f"  Protein plotting length: {protein_len}")

    print("\nGenerating plots...")
    plot_lollipop_multisource_hotspots(
        points=points,
        cluster_summary=cluster_summary,
        protein_len=protein_len,
        out_png=plots_dir / "pyk10_lollipop_multisource_hotspots.png",
        out_pdf=plots_dir / "pyk10_lollipop_multisource_hotspots.pdf",
        shade_min_dataset_support=args.shade_min_dataset_support,
    )

    plot_cluster_matrix(
        cluster_summary=cluster_summary,
        out_png=plots_dir / "pyk10_cluster_evidence_matrix.png",
        out_pdf=plots_dir / "pyk10_cluster_evidence_matrix.pdf",
    )

    write_readme(
        output_dir=args.output_dir,
        sem=sem,
        hunter=hunter,
        meta=meta,
        points=points,
        cluster_summary=cluster_summary,
        cluster_gap=args.cluster_gap,
    )

    print("\nDone.")
    print(f"Output directory: {args.output_dir}")
    print(f"Main plot: {plots_dir / 'pyk10_lollipop_multisource_hotspots.png'}")
    print(f"Cluster matrix: {plots_dir / 'pyk10_cluster_evidence_matrix.png'}")


if __name__ == "__main__":
    main()
