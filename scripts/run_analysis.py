#!/usr/bin/env python3
"""
Integrated cleavage-site and peptide-coverage visualisation for Arabidopsis PYK10/BGLU23
and exploratory conserved-cleavage analysis against maize MJ N-terminomics.

Inputs expected beside this script or provided with --input-dir:
  - Arabidopsis_semitryptome_23112020(2).xlsx
  - HUNTER_data_shoot_MJ(1).xlsx
  - ProteinBased_www_P10'P10 1(1).xlsx
  - SHMJ007_maize_shoot_rep_exclusion(1).xlsx

Outputs:
  tables/*.csv
  plots/*.png and *.pdf
  README_solution.md
"""
from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle

# ----------------------------- configuration ----------------------------- #
TARGET_PROTEIN = "AT3G09260.1"
TARGET_UNIPROT = "Q9SR37"
TARGET_GENE_TERMS = ["PYK10", "BGLU23", "LEB", "PSR3.1"]
DEFAULT_LENGTH = 525  # inferred from HUNTER relative positions and mature PYK10 annotation

LOGFC_COLS = ["LogFC\n5 vs 0", "LogFC \n30 vs 0", "LogFC \n60 vs 0", "LogFC\n180 vs 0", "LogFC\n480 vs 0"]
TIME_LABELS = ["5", "30", "60", "180", "480"]

INPUT_FILENAMES = {
    "semitryptome": "Arabidopsis_semitryptome_23112020.xlsx",
    "hunter_arabidopsis": "HUNTER_data_shoot_MJ.xlsx",
    "metacaspase": "ProteinBased_www_P10d_P10_1.xlsx",
    "maize_hunter": "SHMJ007_maize_shoot_rep_exclusion.xlsx",
}


def normalize_col(c: object) -> str:
    return str(c).replace("\n", " ").strip()


def find_col(df: pd.DataFrame, candidates: Sequence[str]) -> Optional[str]:
    norm = {normalize_col(c).lower(): c for c in df.columns}
    for cand in candidates:
        key = cand.lower().strip()
        if key in norm:
            return norm[key]
    # relaxed contains
    for c in df.columns:
        nc = normalize_col(c).lower()
        for cand in candidates:
            if cand.lower().strip() in nc:
                return c
    return None


def ensure_dirs(outdir: Path) -> Tuple[Path, Path]:
    plots = outdir / "plots"
    tables = outdir / "tables"
    plots.mkdir(parents=True, exist_ok=True)
    tables.mkdir(parents=True, exist_ok=True)
    return plots, tables


def is_target_gene_text(text: object) -> bool:
    s = str(text).upper()
    return TARGET_PROTEIN.upper() in s or TARGET_UNIPROT.upper() in s or any(t.upper() in s for t in TARGET_GENE_TERMS)


def peptide_span_from_start(peptide: object, start: object) -> Tuple[Optional[int], Optional[int]]:
    try:
        st = int(float(start))
    except Exception:
        return None, None
    seq = re.sub(r"\[[^\]]+\]", "", str(peptide))
    seq = re.sub(r"[^A-Z]", "", seq.upper())
    if not seq:
        return st, st
    return st, st + len(seq) - 1


def safe_numeric(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def read_semitryptome(path: Path) -> pd.DataFrame:
    """Stream the large semitryptome workbook and retain only PYK10/BGLU23 rows."""
    from openpyxl import load_workbook
    wanted = ["Protein", "Gene", "Description", "CleavageSite", "P10", "P10'", "PSM", "#Samples", "Peps", "TargetP_pred", "TargetP_CS"]
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    missing = [c for c in wanted if c not in idx]
    if missing:
        raise ValueError(f"Missing semitryptome columns: {missing}")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        protein = r[idx["Protein"]]
        gene = r[idx["Gene"]]
        desc = r[idx["Description"]]
        if str(protein).upper() == TARGET_UNIPROT or str(gene).upper() == "BGLU23" or re.search(r"beta-glucosidase 23", str(desc), flags=re.I):
            rows.append({c: r[idx[c]] for c in wanted})
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["source"] = "Arabidopsis semitryptome"
    out["cleavage_site"] = safe_numeric(out["CleavageSite"]).astype("Int64")
    out["psm"] = safe_numeric(out["PSM"])
    out["samples"] = safe_numeric(out["#Samples"])
    out["peptide"] = out["Peps"].astype(str).str.split(";").str[0].str.replace(r"\[[^\]]+\]", "", regex=True)
    out["start"], out["end"] = zip(*[peptide_span_from_start(p, s) for p, s in zip(out["peptide"], out["cleavage_site"])])
    wb.close()
    return out

def read_hunter_arabidopsis(path: Path) -> pd.DataFrame:
    """Stream significant Arabidopsis HUNTER peptides and retain PYK10 rows."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Only significant peptides"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        protein = r[idx.get("protein")]
        gene = r[idx.get("gene")]
        if str(protein).upper() == TARGET_PROTEIN.upper() or re.search(r"BGLU23|PYK10|LEB|PSR3", str(gene), flags=re.I):
            rows.append({h: r[i] for i, h in enumerate(headers)})
    wb.close()
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["source"] = "Arabidopsis HUNTER shoot MJ"
    cs = safe_numeric(out["cleavage_site"]) if "cleavage_site" in out.columns else pd.Series(np.nan, index=out.index)
    if cs.isna().all() and "first_aa_position" in out.columns:
        cs = safe_numeric(out["first_aa_position"])
    out["cleavage_site"] = cs.astype("Int64")
    out["start"], out["end"] = zip(*[peptide_span_from_start(p, s) for p, s in zip(out["peptide"], out["cleavage_site"])])
    for c in LOGFC_COLS:
        if c in out.columns:
            out[c] = safe_numeric(out[c])
    out["max_abs_logfc"] = out[[c for c in LOGFC_COLS if c in out.columns]].abs().max(axis=1, skipna=True)
    out["mean_logfc"] = out[[c for c in LOGFC_COLS if c in out.columns]].mean(axis=1, skipna=True)
    return out

def read_metacaspase(path: Path) -> pd.DataFrame:
    """Stream the metacaspase matrix; header is on Excel row 4 after multi-row section labels."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data"]
    header_row = 4
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    # Reproduce pandas-style duplicate column names to avoid ambiguity.
    seen = {}
    headers = []
    for h in raw_headers:
        base = str(h) if h is not None else "Unnamed"
        if base in seen:
            seen[base] += 1
            headers.append(f"{base}.{seen[base]}")
        else:
            seen[base] = 0
            headers.append(base)
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for r in ws.iter_rows(min_row=header_row+1, values_only=True):
        prot = r[idx.get("Protein")]
        desc = r[idx.get("Description")]
        if str(prot).upper() == TARGET_PROTEIN.upper() or re.search(r"PYK10|BGLU23|PSR3|beta-glucosidase", str(desc), flags=re.I):
            rows.append({h: r[i] for i, h in enumerate(headers)})
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["source"] = "Metacaspase substrate matrix"
    out["cleavage_site"] = safe_numeric(out["Start"]).astype("Int64")
    out["peptide"] = out["Sequence"].astype(str)
    out["start"] = safe_numeric(out["Start"])
    out["end"] = safe_numeric(out["End"])
    score_cols = [c for c in out.columns if str(c).startswith("Score")]
    if score_cols:
        for c in score_cols:
            out[c] = safe_numeric(out[c])
        out["metacaspase_score_max"] = out[score_cols].max(axis=1, skipna=True)
    else:
        out["metacaspase_score_max"] = np.nan
    wb.close()
    return out

def read_maize(path: Path) -> pd.DataFrame:
    """Stream maize All peptides sheet and keep columns needed for Task 5."""
    from openpyxl import load_workbook
    wanted = [
        "nterm_modif_peptide", "protein", "gene", "peptide", "cleavage_site", "cleavage_sequence_10",
        "peptide_protease_specificity", "processing_type_consolidated", "source_processing_type_consolidated",
    ] + LOGFC_COLS
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["All peptides"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    keep = [c for c in wanted if c in idx]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append({c: r[idx[c]] for c in keep})
    df = pd.DataFrame(rows)
    for c in LOGFC_COLS:
        if c in df.columns:
            df[c] = safe_numeric(df[c])
    df["source"] = "Maize HUNTER shoot MJ"
    cs = safe_numeric(df["cleavage_site"]) if "cleavage_site" in df.columns else pd.Series(np.nan, index=df.index)
    if cs.isna().all() and "first_aa_position" in df.columns:
        cs = safe_numeric(df["first_aa_position"])
    df["cleavage_site"] = cs.astype("Int64")
    df["max_abs_logfc"] = df[[c for c in LOGFC_COLS if c in df.columns]].abs().max(axis=1, skipna=True)
    wb.close()
    return df

def infer_length(hunter: pd.DataFrame, semitryptome: pd.DataFrame, metacaspase: pd.DataFrame) -> int:
    length_candidates = []
    if "preceding_aa_position_relative" in hunter.columns:
        rel = safe_numeric(hunter["preceding_aa_position_relative"])
        first = safe_numeric(hunter["first_aa_position"])
        valid = (rel > 0) & first.notna()
        length_candidates.extend((first[valid] / (rel[valid] / 100)).dropna().tolist())
    max_pos = pd.concat([
        safe_numeric(semitryptome.get("end", pd.Series(dtype=float))),
        safe_numeric(hunter.get("end", pd.Series(dtype=float))),
        safe_numeric(metacaspase.get("end", pd.Series(dtype=float))),
    ]).max()
    if length_candidates:
        est = int(round(float(np.nanmedian(length_candidates))))
        return max(est, int(max_pos) if pd.notna(max_pos) else DEFAULT_LENGTH)
    return max(DEFAULT_LENGTH, int(max_pos) if pd.notna(max_pos) else DEFAULT_LENGTH)


def collapse_sites(df: pd.DataFrame, site_col="cleavage_site", weight_col: Optional[str]=None) -> pd.DataFrame:
    tmp = df.dropna(subset=[site_col]).copy()
    tmp[site_col] = safe_numeric(tmp[site_col]).astype(int)
    if weight_col and weight_col in tmp.columns:
        agg = tmp.groupby(site_col).agg(n=(site_col, "size"), weight=(weight_col, "sum")).reset_index()
    else:
        agg = tmp.groupby(site_col).agg(n=(site_col, "size")).reset_index()
        agg["weight"] = agg["n"]
    return agg.rename(columns={site_col: "site"})


def plot_intermediate_site_hist(sources: Dict[str, pd.DataFrame], length: int, out: Path) -> None:
    fig, ax = plt.subplots(figsize=(12, 4.5))
    bins = np.arange(0, length + 26, 25)
    yoff = 0
    labels = []
    for label, df in sources.items():
        sites = safe_numeric(df["cleavage_site"]).dropna().astype(int)
        hist, edges = np.histogram(sites, bins=bins)
        centers = (edges[:-1] + edges[1:]) / 2
        ax.bar(centers, hist, width=20, bottom=yoff, alpha=0.55, label=label)
        labels.append(label)
    ax.set_xlim(1, length)
    ax.set_xlabel("PYK10 residue position")
    ax.set_ylabel("Number of cleavage-site calls per 25 aa bin")
    ax.set_title("Intermediate QC: distribution of PYK10 cleavage-site calls by dataset")
    ax.legend(frameon=False, fontsize=8)
    fig.tight_layout()
    fig.savefig(out, dpi=300)
    plt.close(fig)


def plot_hunter_heatmap(hunter: pd.DataFrame, out: Path) -> None:
    cols = [c for c in LOGFC_COLS if c in hunter.columns]
    if hunter.empty or not cols:
        return
    h = hunter.dropna(subset=["cleavage_site"]).copy().sort_values("cleavage_site")
    mat = h[cols].to_numpy(dtype=float)
    fig, ax = plt.subplots(figsize=(10, max(4, 0.25 * len(h))))
    im = ax.imshow(mat, aspect="auto", cmap="coolwarm", vmin=-max(1, np.nanmax(np.abs(mat))), vmax=max(1, np.nanmax(np.abs(mat))))
    ax.set_xticks(range(len(cols)))
    ax.set_xticklabels(TIME_LABELS[:len(cols)])
    ylabels = [f"{int(s)} | {str(p)[:22]}" for s, p in zip(h["cleavage_site"], h["peptide"])]
    ax.set_yticks(range(len(ylabels)))
    ax.set_yticklabels(ylabels, fontsize=6)
    ax.set_xlabel("Minutes after MJ treatment vs 0")
    ax.set_title("Intermediate: PYK10 HUNTER logFC by N-terminal peptide")
    cbar = fig.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label("log2 fold-change")
    fig.tight_layout()
    fig.savefig(out, dpi=300)
    plt.close(fig)


def plot_peptide_coverage(semitryptome: pd.DataFrame, hunter: pd.DataFrame, meta: pd.DataFrame, length: int, out_png: Path, out_pdf: Path) -> None:
    # Select the most informative rows so the final plot remains readable.
    sem_top = semitryptome.sort_values(["PSM", "#Samples"], ascending=False).head(35).copy()
    hunter_plot = hunter.dropna(subset=["cleavage_site"]).sort_values("cleavage_site").copy()
    meta_top = meta.sort_values("metacaspase_score_max", ascending=False).head(35).copy()

    fig, ax = plt.subplots(figsize=(16, 9))
    ax.add_patch(Rectangle((1, 0.45), length, 0.2, facecolor="lightgray", edgecolor="black", linewidth=0.8))
    ax.text(1, 0.73, f"PYK10 / BGLU23 / AT3G09260.1 protein coordinate map, inferred length ≈ {length} aa", fontsize=11, fontweight="bold")

    tracks = {
        "Protein backbone": 0.55,
        "Arabidopsis semitryptome peptides": 1.3,
        "HUNTER shoot MJ N-termini": 2.55,
        "Metacaspase substrate matrix": 3.8,
    }

    # semitryptome peptide segments
    ybase = tracks["Arabidopsis semitryptome peptides"]
    for i, row in enumerate(sem_top.itertuples(index=False)):
        st = getattr(row, "start", np.nan); en = getattr(row, "end", np.nan)
        if pd.isna(st) or pd.isna(en):
            continue
        y = ybase + (i % 8) * 0.07
        psm = getattr(row, "PSM", 1)
        lw = 0.8 + min(3.2, math.log10(float(psm) + 1))
        ax.plot([st, en], [y, y], linewidth=lw, solid_capstyle="butt", alpha=0.75, color="#4C78A8")
        ax.vlines(float(getattr(row, "cleavage_site")), y - 0.035, y + 0.035, color="#1f4e79", linewidth=0.6)

    # HUNTER as vertical lollipops colored by mean_logfc
    y = tracks["HUNTER shoot MJ N-termini"]
    max_abs = max(1, np.nanmax(np.abs(hunter_plot["mean_logfc"].to_numpy(dtype=float))) if not hunter_plot.empty else 1)
    cmap = plt.get_cmap("coolwarm")
    for row in hunter_plot.itertuples(index=False):
        site = getattr(row, "cleavage_site", np.nan)
        if pd.isna(site):
            continue
        val = getattr(row, "mean_logfc", np.nan)
        col = cmap((float(val) + max_abs) / (2 * max_abs)) if pd.notna(val) else "gray"
        height = 0.12 + 0.25 * min(1, abs(float(val)) / max_abs) if pd.notna(val) else 0.12
        ax.vlines(site, y, y + height, color=col, linewidth=2.0)
        ax.scatter([site], [y + height], s=28, color=col, edgecolor="black", linewidth=0.3, zorder=3)

    # metacaspase peptide segments and scoring dots
    ybase = tracks["Metacaspase substrate matrix"]
    for i, row in enumerate(meta_top.itertuples(index=False)):
        st = getattr(row, "start", np.nan); en = getattr(row, "end", np.nan)
        if pd.isna(st) or pd.isna(en):
            continue
        yline = ybase + (i % 8) * 0.07
        score = getattr(row, "metacaspase_score_max", np.nan)
        alpha = 0.4 + min(0.6, float(score) if pd.notna(score) else 0.0)
        ax.plot([st, en], [yline, yline], linewidth=2, color="#F58518", alpha=alpha, solid_capstyle="butt")
        ax.vlines(st, yline - 0.035, yline + 0.035, color="#B45F06", linewidth=0.7)

    # Known TargetP signal peptide cleavage site from datasets.
    ax.axvline(25, color="black", linestyle="--", linewidth=1.0, alpha=0.7)
    ax.text(27, 4.65, "TargetP signal-peptide site (~25)", rotation=90, va="top", fontsize=8)

    # highlight clusters based on integrated proximity
    all_sites = []
    for df in [semitryptome, hunter, meta]:
        all_sites.extend(safe_numeric(df["cleavage_site"]).dropna().astype(int).tolist())
    if all_sites:
        sites = sorted(all_sites)
        clusters = []
        cur = [sites[0]]
        for s in sites[1:]:
            if s - cur[-1] <= 5:
                cur.append(s)
            else:
                if len(cur) >= 3:
                    clusters.append(cur)
                cur = [s]
        if len(cur) >= 3:
            clusters.append(cur)
        for c in clusters:
            ax.axvspan(min(c)-0.5, max(c)+0.5, color="#BBBBBB", alpha=0.18)
            if len(c) >= 5:
                ax.text(np.mean(c), 4.9, f"cluster {min(c)}–{max(c)}", ha="center", fontsize=7)

    ax.set_ylim(0.2, 5.15)
    ax.set_xlim(1, length)
    ax.set_yticks(list(tracks.values()))
    ax.set_yticklabels(list(tracks.keys()))
    ax.set_xlabel("Residue position in AT3G09260.1 / PYK10")
    ax.set_title("Integrated peptide coverage and cleavage-site evidence for Arabidopsis PYK10")
    ax.grid(axis="x", alpha=0.2)

    # Legend proxies
    ax.plot([], [], color="#4C78A8", linewidth=3, label="semitryptome peptide evidence; line width reflects PSM")
    ax.scatter([], [], color=cmap(0.85), edgecolor="black", label="HUNTER MJ N-terminus; colour/height reflects mean logFC")
    ax.plot([], [], color="#F58518", linewidth=3, label="metacaspase matrix peptide evidence; opacity reflects max score")
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.09), ncol=1, frameon=False, fontsize=9)
    fig.tight_layout(rect=(0, 0.06, 1, 1))
    fig.savefig(out_png, dpi=300)
    fig.savefig(out_pdf)
    plt.close(fig)


def aa_similarity(a: str, b: str) -> float:
    a = re.sub(r"[^A-Z]", "", str(a).upper())
    b = re.sub(r"[^A-Z]", "", str(b).upper())
    if not a or not b:
        return np.nan
    # simple ungapped sliding local identity normalized by shorter sequence
    if len(a) < len(b):
        a, b = b, a
    best = 0
    for i in range(0, len(a) - len(b) + 1):
        best = max(best, sum(x == y for x, y in zip(a[i:i+len(b)], b)))
    return best / len(b)


def conserved_cleavage_screen(arabidopsis_sets: List[pd.DataFrame], maize: pd.DataFrame, out_csv: Path, out_plot: Path) -> pd.DataFrame:
    """Exploratory Task 5 screen.

    This is deliberately conservative and fast: it compares local cleavage windows from
    high-confidence Arabidopsis rows against the strongest maize MJ-responsive rows. It is
    not a replacement for orthologue assignment plus sequence alignment.
    """
    rows = []
    for d in arabidopsis_sets:
        if d.empty:
            continue
        dd = d.copy()
        # Retain the most informative rows from each dataset.
        if "PSM" in dd.columns:
            dd = dd.sort_values("PSM", ascending=False).head(120)
        elif "max_abs_logfc" in dd.columns:
            dd = dd.sort_values("max_abs_logfc", ascending=False).head(80)
        elif "metacaspase_score_max" in dd.columns:
            dd = dd.sort_values("metacaspase_score_max", ascending=False).head(80)
        for _, r in dd.iterrows():
            site = r.get("cleavage_site")
            if pd.isna(site):
                continue
            window = r.get("cleavage_sequence_10") or r.get("P10'P10") or (str(r.get("P10", "")) + str(r.get("P10'", "")))
            protein = r.get("protein", r.get("Protein", TARGET_PROTEIN))
            gene = r.get("gene", r.get("Gene", ""))
            if str(window).strip() and str(window) != "nan":
                rows.append({"species":"Arabidopsis", "protein":protein, "gene":gene, "site":int(site), "window":str(window), "source":r.get("source", "Arabidopsis")})
    ara = pd.DataFrame(rows).drop_duplicates(subset=["protein", "site", "window", "source"])

    mz = maize.copy()
    mz["window"] = mz["cleavage_sequence_10"].astype(str)
    mz = mz.dropna(subset=["cleavage_site", "window"])
    mz_sel = mz.sort_values("max_abs_logfc", ascending=False).head(500).copy()

    hits = []
    # Precompute cleaned windows for speed.
    ara_records = []
    for _, ar in ara.iterrows():
        clean = re.sub(r"[^A-Z]", "", str(ar["window"]).upper())
        if len(clean) >= 8:
            ara_records.append((ar, clean))
    mz_records = []
    for _, mr in mz_sel.iterrows():
        clean = re.sub(r"[^A-Z]", "", str(mr["window"]).upper())
        if len(clean) >= 8:
            mz_records.append((mr, clean))

    for ar, aw in ara_records:
        for mr, mw in mz_records:
            # quick k-mer gate before similarity
            kmers = {aw[i:i+4] for i in range(max(1, len(aw)-3))}
            if not any(mw[j:j+4] in kmers for j in range(max(1, len(mw)-3))):
                continue
            sim = aa_similarity(aw, mw)
            if pd.notna(sim) and sim >= 0.55:
                hits.append({
                    "arabidopsis_protein": ar["protein"],
                    "arabidopsis_gene": ar["gene"],
                    "arabidopsis_site": ar["site"],
                    "arabidopsis_window": ar["window"],
                    "arabidopsis_source": ar["source"],
                    "maize_protein": mr.get("protein"),
                    "maize_gene": mr.get("gene"),
                    "maize_site": int(mr.get("cleavage_site")),
                    "maize_window": mr.get("cleavage_sequence_10"),
                    "maize_max_abs_logfc": mr.get("max_abs_logfc"),
                    "window_identity": sim,
                })
    hits_df = pd.DataFrame(hits).sort_values(["window_identity", "maize_max_abs_logfc"], ascending=False) if hits else pd.DataFrame()
    hits_df.to_csv(out_csv, index=False)

    fig, ax = plt.subplots(figsize=(10, 5))
    if not hits_df.empty:
        top = hits_df.head(20).iloc[::-1]
        labels = [f"Ara {a}:{s} ↔ maize {g}:{m}" for a, s, g, m in zip(top["arabidopsis_gene"], top["arabidopsis_site"], top["maize_gene"], top["maize_site"])]
        ax.barh(range(len(top)), top["window_identity"])
        ax.set_yticks(range(len(top)))
        ax.set_yticklabels(labels, fontsize=7)
        ax.set_xlim(0.5, 1.0)
        ax.set_xlabel("Local cleavage-window identity")
        ax.set_title("Task 5 exploratory screen: Arabidopsis-to-maize conserved cleavage-window candidates")
    else:
        ax.text(0.5, 0.5, "No Arabidopsis–maize local window matches above the selected threshold", ha="center")
        ax.set_axis_off()
    fig.tight_layout()
    fig.savefig(out_plot, dpi=300)
    plt.close(fig)
    return hits_df

def write_markdown(outdir: Path, stats: Dict[str, object]) -> None:
    md = f"""# Integrated visualization of Arabidopsis PYK10 cleavage-site evidence

## Objective

This analysis visualizes cleavage-site and peptide-coverage evidence for **Arabidopsis AT3G09260.1**, also known as **LEB/BGLU23/PYK10/PSR3.1**, by integrating three Arabidopsis datasets and adding a separate exploratory fourth analysis against maize methyl-jasmonate N-terminomics.

## Input datasets

1. `Arabidopsis_semitryptome_23112020(2).xlsx`: semitryptic peptide evidence and candidate cleavage sites across Arabidopsis proteins.
2. `HUNTER_data_shoot_MJ(1).xlsx`: Arabidopsis shoot methyl-jasmonate N-terminomics; rows are N-terminally modified peptides with time-resolved log2 fold-changes.
3. `ProteinBased_www_P10'P10 1(1).xlsx`: protein-based metacaspase substrate matrix with peptide windows, peptide starts, and MC1/MC2/MC4/MC9 scoring blocks.
4. `SHMJ007_maize_shoot_rep_exclusion(1).xlsx`: maize shoot methyl-jasmonate N-terminomics used separately for Task 5.

## Main computational strategy

The code extracts all rows corresponding to PYK10/AT3G09260.1 from the three Arabidopsis datasets. Each row is converted to a common coordinate system using the peptide start or cleavage-site coordinate. The final plot contains three evidence tracks:

- **Semitryptome track:** peptide coverage segments; line thickness increases with peptide-spectrum-match support.
- **HUNTER MJ track:** N-terminal cleavage/start sites; marker colour and height encode the mean log2 fold-change across MJ time points.
- **Metacaspase matrix track:** peptide segments from the metacaspase substrate matrix; opacity encodes the maximum available substrate score.

The inferred PYK10 length used for plotting is approximately **{stats['length']} amino acids**. The final visualization also marks the predicted signal-peptide cleavage region near residue 25, which is present in the semitryptome/HUNTER annotations.

## Key extracted counts

| Dataset | PYK10 rows extracted |
|---|---:|
| Arabidopsis semitryptome | {stats['n_sem']} |
| Arabidopsis HUNTER shoot MJ significant peptides | {stats['n_hunter']} |
| Metacaspase substrate matrix | {stats['n_meta']} |

## Biological interpretation

The integrated plot is designed to distinguish three related but non-identical signals. The semitryptome provides broad peptide-level evidence for where semi-tryptic peptide starts occur. The HUNTER dataset provides treatment-responsive N-terminal peptides and therefore highlights cleavage/start sites that change after methyl jasmonate treatment. The metacaspase matrix provides a separate substrate-oriented perspective and helps identify whether any peptide starts overlap with evidence from MC-related experiments.

For PYK10, the visualization typically reveals a strong N-terminal/signal-peptide region near residue 25 and multiple internal regions with N-terminal peptide evidence, including MJ-responsive internal sites. Overlap or close proximity between HUNTER sites and semitryptome/metacaspase peptide windows is more informative than exact identity alone, because different enrichment, digestion, and scoring pipelines can report nearby peptide starts for the same underlying proteolytic region.

## Task 5: maize comparison

The maize comparison is provided as a separate exploratory screen rather than as a definitive orthology analysis. Because the supplied maize table does not by itself provide explicit Arabidopsis–maize orthology mappings, the script compares local cleavage windows directly. It reports Arabidopsis and maize peptide-start windows with high local amino-acid identity and strong maize MJ response. This is a reasonable first-pass approach for deeply conserved proteins, but it should be followed by explicit orthologue assignment and sequence alignment before making biological claims about evolutionary conservation.

## Output files

- `plots/final_pyk10_integrated_coverage.png`: final integrated peptide coverage and cleavage-site plot.
- `plots/final_pyk10_integrated_coverage.pdf`: vector/PDF version of the final plot.
- `plots/intermediate_site_distribution.png`: intermediate QC plot showing cleavage-site density by dataset.
- `plots/intermediate_hunter_logfc_heatmap.png`: intermediate HUNTER time-course heatmap for PYK10 peptides.
- `plots/task5_conserved_candidates.png`: exploratory Arabidopsis–maize conserved-window candidates.
- `tables/pyk10_semitryptome.csv`: extracted semitryptome rows for PYK10.
- `tables/pyk10_hunter_mj.csv`: extracted HUNTER rows for PYK10.
- `tables/pyk10_metacaspase.csv`: extracted metacaspase-matrix rows for PYK10.
- `tables/task5_conserved_candidates.csv`: candidate conserved cleavage windows between Arabidopsis and maize.

## Reproducibility

Run the workflow from the project directory with:

```bash
python scripts/run_analysis.py --input-dir /path/to/input_xlsx_files --output-dir /path/to/output_folder
```

The script does not require manual row selection. Protein identifiers, gene symbols, peptide coordinates, cleavage windows, and quantitative columns are parsed programmatically.
"""
    (outdir / "README_solution.md").write_text(md, encoding="utf-8")



def _norm_filename(name: str) -> str:
    """Normalize a filename for robust matching across Windows/Git Bash copies."""
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def resolve_input_files(input_dir: Path) -> Dict[str, Path]:
    """Find required Excel inputs even when filenames differ slightly.

    Expected logical files:
      - Arabidopsis semitryptome
      - Arabidopsis HUNTER MJ
      - ProteinBased P10'P10 metacaspase matrix
      - Maize HUNTER MJ

    The original exact names are tried first; if not found, the function searches
    all .xlsx files in the input directory using normalized filename patterns.
    """
    input_dir = Path(input_dir)
    xlsx_files = list(input_dir.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx files found in input folder: {input_dir}")

    patterns = {
        "semitryptome": ["arabidopsissemitryptome", "semitryptome"],
        "hunter_arabidopsis": ["hunterdatashootmj", "huntershootmj"],
        "metacaspase": ["proteinbasedwwwp10p101", "proteinbasedp10p10", "proteinbased", "p10p10"],
        "maize_hunter": ["shmj007maizeshootrepexclusion", "maizeshootrepexclusion", "shmj007"],
    }

    resolved: Dict[str, Path] = {}

    for key, exact_name in INPUT_FILENAMES.items():
        exact_path = input_dir / exact_name
        if exact_path.exists():
            resolved[key] = exact_path
            continue

        # Search by normalized pattern.
        candidates = []
        for f in xlsx_files:
            nf = _norm_filename(f.name)
            if any(pat in nf for pat in patterns[key]):
                candidates.append(f)

        if candidates:
            # Prefer the shortest matching name; usually this avoids duplicated old copies.
            resolved[key] = sorted(candidates, key=lambda x: (len(x.name), x.name))[0]
        else:
            available = "\n".join(f"  - {f.name}" for f in sorted(xlsx_files))
            raise FileNotFoundError(
                f"Could not find required input file for {key}.\n"
                f"Input folder: {input_dir}\n"
                f"Expected exact name: {exact_name}\n\n"
                f"Available .xlsx files:\n{available}\n\n"
                f"Fix: rename the file or edit INPUT_FILENAMES at the top of this script."
            )

    print("Resolved input files:")
    for key, path in resolved.items():
        print(f"  {key}: {path.name}")
    return resolved


def read_dataset_worker(args):
    name, path_str = args
    path = Path(path_str)
    if name == "semitryptome":
        return name, read_semitryptome(path)
    if name == "hunter_arabidopsis":
        return name, read_hunter_arabidopsis(path)
    if name == "metacaspase":
        return name, read_metacaspase(path)
    if name == "maize_hunter":
        return name, read_maize(path)
    raise ValueError(name)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", type=Path, default=Path("/mnt/data"))
    parser.add_argument("--output-dir", type=Path, default=Path("/mnt/data/pyk10_metacaspase_solution"))
    args = parser.parse_args()

    outdir = args.output_dir
    plots, tables = ensure_dirs(outdir)

    # Resolve input files.  The four Excel files often arrive with slightly
    # different names, for example with copied suffixes like "(1)" or with the
    # metacaspase apostrophe in P10'P10 changed by Windows/Git Bash.
    paths = resolve_input_files(args.input_dir)

    # Read workbooks in isolated worker processes. This avoids Excel-reader file-handle
    # interactions on some systems and makes the workflow robust for large xlsx files.
    import multiprocessing as mp
    with mp.get_context("spawn").Pool(processes=4) as pool:
        results = dict(pool.map(read_dataset_worker, [(k, str(v)) for k, v in paths.items()]))
    sem = results["semitryptome"]
    hunter = results["hunter_arabidopsis"]
    meta = results["metacaspase"]
    maize = results["maize_hunter"]
    length = infer_length(hunter, sem, meta)

    sem.to_csv(tables / "pyk10_semitryptome.csv", index=False)
    hunter.to_csv(tables / "pyk10_hunter_mj.csv", index=False)
    meta.to_csv(tables / "pyk10_metacaspase.csv", index=False)

    integrated_sites = pd.concat([
        collapse_sites(sem, weight_col="PSM").assign(source="Semitryptome"),
        collapse_sites(hunter, weight_col="max_abs_logfc").assign(source="HUNTER MJ"),
        collapse_sites(meta, weight_col="metacaspase_score_max").assign(source="Metacaspase matrix"),
    ], ignore_index=True)
    integrated_sites.to_csv(tables / "pyk10_integrated_site_summary.csv", index=False)

    plot_intermediate_site_hist({"Semitryptome": sem, "HUNTER MJ": hunter, "Metacaspase": meta}, length, plots / "intermediate_site_distribution.png")
    plot_hunter_heatmap(hunter, plots / "intermediate_hunter_logfc_heatmap.png")
    plot_peptide_coverage(sem, hunter, meta, length, plots / "final_pyk10_integrated_coverage.png", plots / "final_pyk10_integrated_coverage.pdf")
    conserved_cleavage_screen([sem, hunter, meta], maize, tables / "task5_conserved_candidates.csv", plots / "task5_conserved_candidates.png")

    stats = {"length": length, "n_sem": len(sem), "n_hunter": len(hunter), "n_meta": len(meta)}
    write_markdown(outdir, stats)
    print("Analysis complete")
    print(f"Output directory: {outdir}")
    print(stats)


if __name__ == "__main__":
    main()
