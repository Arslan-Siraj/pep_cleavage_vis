#!/usr/bin/env python3
"""Exploratory conserved cleavage-window screen for Task 5.

This script compares Arabidopsis and maize methyl-jasmonate N-terminomics tables without
requiring external orthology resources. It searches for highly similar local cleavage
windows and exact/near-exact N-terminal peptide sequence matches. Candidate hits should be
interpreted as hypotheses and followed by explicit orthologue assignment/alignment.
"""
from __future__ import annotations
import argparse, re, math
from pathlib import Path
from collections import defaultdict
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

LOGFC_COLS = ["LogFC\n5 vs 0", "LogFC \n30 vs 0", "LogFC \n60 vs 0", "LogFC\n180 vs 0", "LogFC\n480 vs 0"]

def clean(s):
    return re.sub(r"[^A-Z]", "", str(s).upper())

def safe_num(x):
    return pd.to_numeric(x, errors='coerce')

def read_hunter_all(path: Path, sheet: str, species: str) -> pd.DataFrame:
    wanted = ["nterm_modif_peptide","protein","gene","peptide","cleavage_site","cleavage_sequence_10","first_aa_position","peptide_protease_specificity","processing_type_consolidated","source_processing_type_consolidated"] + LOGFC_COLS
    wb=load_workbook(path, read_only=True, data_only=True)
    ws=wb[sheet]
    headers=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
    idx={h:i for i,h in enumerate(headers)}
    keep=[c for c in wanted if c in idx]
    rows=[]
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append({c:r[idx[c]] for c in keep})
    wb.close()
    df=pd.DataFrame(rows)
    df['species']=species
    # Some supplied sheets store sequence text in cleavage_site and numeric coordinate in first_aa_position.
    cs=safe_num(df['cleavage_site']) if 'cleavage_site' in df else pd.Series(np.nan,index=df.index)
    if cs.isna().all() and 'first_aa_position' in df:
        cs=safe_num(df['first_aa_position'])
    df['cleavage_site_numeric']=cs
    for c in LOGFC_COLS:
        if c in df.columns: df[c]=safe_num(df[c])
    df['max_abs_logfc']=df[[c for c in LOGFC_COLS if c in df.columns]].abs().max(axis=1, skipna=True)
    df['clean_window']=df.get('cleavage_sequence_10', pd.Series('', index=df.index)).map(clean)
    df['clean_peptide']=df.get('peptide', pd.Series('', index=df.index)).map(clean)
    return df

def local_identity(a,b):
    a=clean(a); b=clean(b)
    if len(a)<8 or len(b)<8: return np.nan
    if len(a)<len(b): a,b=b,a
    best=0
    L=len(b)
    for i in range(len(a)-L+1):
        best=max(best, sum(x==y for x,y in zip(a[i:i+L], b)))
    return best/L

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--input-dir', type=Path, default=Path('/mnt/data'))
    ap.add_argument('--output-dir', type=Path, default=Path('/mnt/data/pyk10_metacaspase_solution'))
    args=ap.parse_args()
    tables=args.output_dir/'tables'; plots=args.output_dir/'plots'; tables.mkdir(exist_ok=True,parents=True); plots.mkdir(exist_ok=True,parents=True)
    ara=read_hunter_all(args.input_dir/'HUNTER_data_shoot_MJ(1).xlsx','All peptides','Arabidopsis')
    maize=read_hunter_all(args.input_dir/'SHMJ007_maize_shoot_rep_exclusion(1).xlsx','All peptides','Maize')
    # Keep rows likely to be interpretable: long enough windows/peptides, strong response, and canonical conserved protein hints.
    conserved_terms=re.compile(r'PSB|PSA|RIBOSOM|HISTONE|ATP|CHLOROPHYLL|OXYGEN|NUCLEOSOME|CYTOCHROME|TUBULIN|ACTIN', re.I)
    ara_pool=ara[(ara.clean_window.str.len()>=12) & (ara.max_abs_logfc.fillna(0)>=1)].copy()
    maize_pool=maize[(maize.clean_window.str.len()>=12) & (maize.max_abs_logfc.fillna(0)>=1)].copy()
    # include conserved-term rows even with weaker logFC
    ara_pool=pd.concat([ara_pool, ara[ara.gene.astype(str).str.contains(conserved_terms, na=False) | ara.protein.astype(str).str.contains(conserved_terms, na=False)]], ignore_index=True).drop_duplicates()
    maize_pool=pd.concat([maize_pool, maize[maize.gene.astype(str).str.contains(conserved_terms, na=False) | maize.protein.astype(str).str.contains(conserved_terms, na=False)]], ignore_index=True).drop_duplicates()
    # restrict to top responses for tractability but keep enough to catch chloroplast/ribosomal proteins
    ara_pool=ara_pool.sort_values('max_abs_logfc', ascending=False).head(2500)
    maize_pool=maize_pool.sort_values('max_abs_logfc', ascending=False).head(2500)
    # kmer index on maize windows and peptides
    idx=defaultdict(list)
    for mi,m in maize_pool.iterrows():
        seqs=[m.clean_window, m.clean_peptide]
        for seq in seqs:
            if len(seq)>=6:
                for j in range(len(seq)-5): idx[seq[j:j+6]].append(mi)
    hits=[]
    for ai,a in ara_pool.iterrows():
        candidates=set()
        for seq in [a.clean_window, a.clean_peptide]:
            if len(seq)>=6:
                for j in range(len(seq)-5): candidates.update(idx.get(seq[j:j+6], []))
        for mi in candidates:
            m=maize_pool.loc[mi]
            win_id=local_identity(a.clean_window, m.clean_window)
            pep_id=local_identity(a.clean_peptide, m.clean_peptide)
            best=np.nanmax([win_id if pd.notna(win_id) else np.nan, pep_id if pd.notna(pep_id) else np.nan])
            if pd.notna(best) and best>=0.65:
                hits.append({
                    'arabidopsis_protein':a.get('protein'), 'arabidopsis_gene':a.get('gene'), 'arabidopsis_site':a.get('cleavage_site_numeric'),
                    'arabidopsis_peptide':a.get('peptide'), 'arabidopsis_window':a.get('cleavage_sequence_10'), 'arabidopsis_max_abs_logfc':a.get('max_abs_logfc'),
                    'maize_protein':m.get('protein'), 'maize_gene':m.get('gene'), 'maize_site':m.get('cleavage_site_numeric'),
                    'maize_peptide':m.get('peptide'), 'maize_window':m.get('cleavage_sequence_10'), 'maize_max_abs_logfc':m.get('max_abs_logfc'),
                    'window_identity':win_id, 'peptide_identity':pep_id, 'best_identity':best,
                })
    hits=pd.DataFrame(hits)
    if not hits.empty:
        hits=hits.sort_values(['best_identity','arabidopsis_max_abs_logfc','maize_max_abs_logfc'], ascending=False).drop_duplicates()
    hits.to_csv(tables/'task5_conserved_hunter_cross_species_candidates.csv', index=False)
    # plot top
    fig,ax=plt.subplots(figsize=(11,6))
    if hits.empty:
        ax.text(.5,.5,'No candidates above threshold',ha='center'); ax.set_axis_off()
    else:
        top=hits.head(25).iloc[::-1]
        labels=[f"{ag}:{int(asite) if pd.notna(asite) else '?'} ↔ {mg}:{int(msite) if pd.notna(msite) else '?'}" for ag,asite,mg,msite in zip(top.arabidopsis_gene,top.arabidopsis_site,top.maize_gene,top.maize_site)]
        ax.barh(range(len(top)), top.best_identity)
        ax.set_yticks(range(len(top))); ax.set_yticklabels(labels,fontsize=7)
        ax.set_xlim(0.6,1.0); ax.set_xlabel('Best local identity: cleavage window or peptide')
        ax.set_title('Task 5: exploratory conserved Arabidopsis–maize cleavage candidates')
    fig.tight_layout(); fig.savefig(plots/'task5_conserved_hunter_cross_species_candidates.png', dpi=300); plt.close(fig)
    print(f'Arabidopsis pool: {len(ara_pool)}; maize pool: {len(maize_pool)}; hits: {len(hits)}')

if __name__=='__main__':
    main()
