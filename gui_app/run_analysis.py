#!/usr/bin/env python3
"""
run_analysis_lollipop.py

Presentation-friendly lollipop visualization for Arabidopsis AT3G09260.1
(PYK10 / BGLU23 / LEB / PSR3.1) cleavage-site evidence.

This script reads the raw Excel files, extracts PYK10 rows from three Arabidopsis
datasets, converts them to a common protein-coordinate system, clusters nearby
cleavage sites, and generates a clean lollipop plot.

Main input files expected in --input-dir:
  1. Arabidopsis semitryptome Excel file
  2. HUNTER_data_shoot_MJ Excel file
  3. ProteinBased_www_P10'P10 Excel file

The script is robust to copied/renamed files such as:
  Arabidopsis_semitryptome_23112020.xlsx
  Arabidopsis_semitryptome_23112020(2).xlsx
  HUNTER_data_shoot_MJ.xlsx
  HUNTER_data_shoot_MJ(1).xlsx
  ProteinBased_www_P10'P10 1.xlsx
  ProteinBased_www_P10'P10 1(1).xlsx
  ProteinBased_www_P10d_P10_1.xlsx

Usage:
  python scripts/run_analysis_lollipop.py --input-dir data --output-dir output

Outputs:
  output/tables/pyk10_semitryptome.csv
  output/tables/pyk10_hunter_mj.csv
  output/tables/pyk10_metacaspase.csv
  output/tables/pyk10_lollipop_clusters.csv
  output/tables/pyk10_lollipop_points.csv
  output/plots/pyk10_lollipop_multisource_hotspots.png
  output/plots/pyk10_lollipop_multisource_hotspots.pdf
  output/README_lollipop.md
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------

DEFAULT_TARGET = "AT3G09260.1"
DEFAULT_ALIASES = ["Q9SR37", "PYK10", "BGLU23", "LEB", "PSR3.1"]
DEFAULT_LENGTH = 525

LOGFC_COLS = [
    "LogFC\n5 vs 0",
    "LogFC \n30 vs 0",
    "LogFC \n60 vs 0",
    "LogFC\n180 vs 0",
    "LogFC\n480 vs 0",
]

# These are preferred names, but the resolver below also searches flexibly.
INPUT_FILENAMES = {
    "semitryptome": "Arabidopsis_semitryptome_23112020.xlsx",
    "hunter_arabidopsis": "HUNTER_data_shoot_MJ.xlsx",
    "metacaspase": "ProteinBased_www_P10d_P10_1.xlsx",
}


# ---------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------

def safe_numeric(x) -> pd.Series:
    return pd.to_numeric(x, errors="coerce")


def clean_peptide_sequence(seq: object) -> str:
    """Remove modification notation and non-amino-acid characters."""
    seq = re.sub(r"\[[^\]]+\]", "", str(seq))
    return re.sub(r"[^A-Z]", "", seq.upper())


def peptide_span_from_start(peptide: object, start: object) -> Tuple[Optional[int], Optional[int]]:
    """Calculate peptide start/end coordinates from start coordinate and sequence."""
    try:
        st = int(float(start))
    except Exception:
        return None, None

    seq = clean_peptide_sequence(peptide)
    if not seq:
        return st, st

    return st, st + len(seq) - 1


def normalize_filename(name: str) -> str:
    """
    Normalize filename for robust matching.
    Example:
      ProteinBased_www_P10'P10 1(1).xlsx
      ProteinBased_www_P10d_P10_1.xlsx
    both normalize to similar alphanumeric strings.
    """
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def resolve_input_files(input_dir: Path) -> Dict[str, Path]:
    """
    Resolve input Excel files robustly.

    This avoids failures caused by apostrophes, spaces, copied suffixes such as (1),
    and sanitized filenames such as P10d_P10.
    """
    input_dir = Path(input_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")

    xlsx_files = sorted(input_dir.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx files found in input directory: {input_dir}")

    patterns = {
        "semitryptome": [
            "arabidopsissemitryptome",
            "semitryptome",
        ],
        "hunter_arabidopsis": [
            "hunterdatashootmj",
            "huntershootmj",
            "shootmj",
        ],
        "metacaspase": [
            "proteinbasedwwwp10p101",
            "proteinbasedp10p10",
            "proteinbasedwwwp10dp101",
            "proteinbased",
            "p10p10",
            "p10dp10",
        ],
    }

    resolved: Dict[str, Path] = {}

    for key, preferred in INPUT_FILENAMES.items():
        preferred_path = input_dir / preferred
        if preferred_path.exists():
            resolved[key] = preferred_path
            continue

        candidates = []
        for f in xlsx_files:
            nf = normalize_filename(f.name)
            if any(pat in nf for pat in patterns[key]):
                candidates.append(f)

        if not candidates:
            available = "\n".join(f"  - {f.name}" for f in xlsx_files)
            raise FileNotFoundError(
                f"Could not find required Excel file for '{key}'.\n"
                f"Input directory: {input_dir}\n"
                f"Preferred filename: {preferred}\n\n"
                f"Available .xlsx files:\n{available}\n\n"
                f"Fix: rename the file, or edit INPUT_FILENAMES in this script."
            )

        # Prefer the shortest matching filename, then alphabetic order.
        resolved[key] = sorted(candidates, key=lambda p: (len(p.name), p.name.lower()))[0]

    print("Resolved input files:")
    for key, path in resolved.items():
        print(f"  {key}: {path.name}")

    return resolved


def ensure_output_dirs(output_dir: Path) -> Tuple[Path, Path]:
    output_dir = Path(output_dir)
    plots = output_dir / "plots"
    tables = output_dir / "tables"
    plots.mkdir(parents=True, exist_ok=True)
    tables.mkdir(parents=True, exist_ok=True)
    return plots, tables


def unique_headers(raw_headers: Sequence[object]) -> List[str]:
    """
    Make duplicate Excel headers unique, similar to pandas behavior.
    """
    seen = {}
    headers = []
    for h in raw_headers:
        base = str(h) if h is not None else "Unnamed"
        if base in seen:
            seen[base] += 1
            headers.append(f"{base}.{seen[base]}")
        else:
            seen[base] = 0
            headers.append(base)
    return headers


@dataclass(frozen=True)
class TargetSpec:
    """Protein identifier bundle used to match rows across heterogeneous tables."""
    query: str
    display_name: str
    output_prefix: str
    uniprot_id: str = ""
    aliases: Tuple[str, ...] = ()
    protein_length: Optional[int] = None

    @property
    def terms(self) -> Tuple[str, ...]:
        vals = [self.query, self.uniprot_id, *self.aliases]
        out = []
        seen = set()
        for v in vals:
            v = str(v).strip()
            if not v or v.lower() == "nan":
                continue
            key = normalize_match_text(v)
            if key and key not in seen:
                seen.add(key)
                out.append(v)
        return tuple(out)


def normalize_match_text(value: object) -> str:
    """Normalize identifiers while preserving enough specificity for matching."""
    return re.sub(r"[^A-Z0-9]+", "", str(value).upper())


def split_alias_string(value: object) -> List[str]:
    """Split comma/semicolon/pipe-delimited alias strings from CLI or mapping files."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    return [x.strip() for x in re.split(r"[,;|]", str(value)) if x.strip()]


def make_output_prefix(value: str) -> str:
    """Create a safe filename prefix from the requested target name."""
    prefix = re.sub(r"[^A-Za-z0-9]+", "_", str(value).strip()).strip("_").lower()
    return prefix or "target_protein"


def target_matches_fields(target: TargetSpec, fields: Sequence[object], allow_substring: bool = True) -> bool:
    """
    Return True if any target term appears in the supplied row fields.

    Exact normalized matching handles identifiers such as AT3G09260.1 and Q9SR37.
    Substring matching is retained for descriptive fields, so gene symbols and
    protein names such as PYK10 or beta-glucosidase 23 can also be found.
    """
    row_text = " ".join("" if v is None else str(v) for v in fields)
    norm_row = normalize_match_text(row_text)
    row_upper = row_text.upper()

    for term in target.terms:
        norm_term = normalize_match_text(term)
        if not norm_term:
            continue

        # Identifier-like exact/token matching after normalization.
        if norm_term == norm_row:
            return True
        if re.search(rf"(?<![A-Z0-9]){re.escape(norm_term)}(?![A-Z0-9])", norm_row):
            return True

        # Descriptive substring matching. This is useful for names in Description.
        if allow_substring and len(str(term).strip()) >= 3 and str(term).upper() in row_upper:
            return True

    return False


def load_alias_map(alias_map: Optional[Path], query: str) -> Tuple[List[str], Optional[int], Optional[str], Optional[str]]:
    """
    Optionally load aliases from a CSV/TSV mapping table.

    Supported columns are flexible. Recommended columns:
      target, uniprot_id, aliases, length, display_name, output_prefix

    A row is selected if `query` matches the target field or any alias.
    """
    if alias_map is None:
        return [], None, None, None

    alias_map = Path(alias_map)
    if not alias_map.exists():
        raise FileNotFoundError(f"Alias map does not exist: {alias_map}")

    sep = "\t" if alias_map.suffix.lower() in {".tsv", ".tab"} else ","
    table = pd.read_csv(alias_map, sep=sep)
    lower_cols = {str(c).lower().strip(): c for c in table.columns}

    target_col = next((lower_cols[c] for c in ["target", "query", "protein", "protein_id", "primary_id"] if c in lower_cols), None)
    uniprot_col = next((lower_cols[c] for c in ["uniprot", "uniprot_id", "uniprot_accession", "accession"] if c in lower_cols), None)
    alias_col = next((lower_cols[c] for c in ["aliases", "alias", "gene_terms", "names", "symbols"] if c in lower_cols), None)
    length_col = next((lower_cols[c] for c in ["length", "protein_length", "aa_length"] if c in lower_cols), None)
    display_col = next((lower_cols[c] for c in ["display_name", "label", "name"] if c in lower_cols), None)
    prefix_col = next((lower_cols[c] for c in ["output_prefix", "prefix", "slug"] if c in lower_cols), None)

    query_norm = normalize_match_text(query)
    for _, row in table.iterrows():
        row_terms = []
        if target_col is not None:
            row_terms.append(row.get(target_col))
        if uniprot_col is not None:
            row_terms.append(row.get(uniprot_col))
        if alias_col is not None:
            row_terms.extend(split_alias_string(row.get(alias_col)))

        if any(normalize_match_text(t) == query_norm for t in row_terms if str(t).strip() and str(t).lower() != "nan"):
            aliases = split_alias_string(row.get(alias_col)) if alias_col is not None else []
            if target_col is not None and str(row.get(target_col)).strip() and str(row.get(target_col)).lower() != "nan":
                aliases.append(str(row.get(target_col)).strip())

            length = None
            if length_col is not None and pd.notna(row.get(length_col)):
                length = int(float(row.get(length_col)))

            display = None
            if display_col is not None and pd.notna(row.get(display_col)):
                display = str(row.get(display_col)).strip()

            prefix = None
            if prefix_col is not None and pd.notna(row.get(prefix_col)):
                prefix = str(row.get(prefix_col)).strip()

            return aliases, length, display, prefix

    print(f"WARNING: target '{query}' was not found in alias map {alias_map.name}; using CLI aliases only.")
    return [], None, None, None


def make_target_spec(
    query: str,
    uniprot_id: Optional[str] = None,
    aliases: Optional[Sequence[str]] = None,
    alias_map: Optional[Path] = None,
    protein_length: Optional[int] = None,
    output_prefix: Optional[str] = None,
) -> TargetSpec:
    """Build the target specification from CLI input and an optional alias map.

    The UniProt accession is accepted as a separate field because the semitryptome
    table commonly stores UniProt accessions in its Protein column, whereas HUNTER
    and the metacaspase matrix often use TAIR-style identifiers. The accession is
    still included among search terms, but it is not mixed into the user-facing
    alias box.
    """
    query = str(query).strip()
    uniprot_id = str(uniprot_id).strip() if uniprot_id is not None and str(uniprot_id).strip() else ""
    map_aliases, map_length, map_display, map_prefix = load_alias_map(alias_map, query)

    merged_aliases: List[str] = []
    for source in [map_aliases, aliases or []]:
        for value in source:
            for alias in split_alias_string(value):
                if normalize_match_text(alias) not in {normalize_match_text(query), normalize_match_text(uniprot_id)}:
                    merged_aliases.append(alias)

    if normalize_match_text(query) == normalize_match_text(DEFAULT_TARGET):
        if not uniprot_id:
            uniprot_id = DEFAULT_ALIASES[0]
        merged_aliases.extend(DEFAULT_ALIASES[1:])

    # Deduplicate while preserving order.
    deduped = []
    seen = set()
    for alias in merged_aliases:
        key = normalize_match_text(alias)
        if key and key not in seen:
            seen.add(key)
            deduped.append(alias)

    length = protein_length if protein_length is not None else map_length
    display = map_display or " / ".join([x for x in [query, uniprot_id, *deduped[:5]] if x])
    prefix = output_prefix or map_prefix or make_output_prefix(query)

    return TargetSpec(
        query=query,
        uniprot_id=uniprot_id,
        aliases=tuple(deduped),
        display_name=display,
        output_prefix=make_output_prefix(prefix),
        protein_length=length,
    )


# ---------------------------------------------------------------------
# Excel readers
# ---------------------------------------------------------------------

# dot size = PSM
def read_semitryptome(path: Path, target: TargetSpec) -> pd.DataFrame:
    """
    Read the Arabidopsis semitryptome file and retain rows matching the target protein.

    Expected relevant columns:
      Protein, Gene, Description, CleavageSite, P10, P10', PSM, #Samples, Peps
    """
    wanted = [
        "Protein", "Gene", "Description", "Representative", "CleavageSite",
        "P10", "P10'", "PSM", "PSM_unique", "#Peps", "Peps", "#Samples",
        "TargetP_pred", "TargetP_CS", "TargetP_prob"
    ]

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = unique_headers(raw_headers)
    idx = {h: i for i, h in enumerate(headers)}

    missing = [c for c in ["Protein", "Gene", "Description", "CleavageSite", "PSM", "Peps"] if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Missing semitryptome columns in {path.name}: {missing}")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        protein = r[idx["Protein"]]
        gene = r[idx["Gene"]]
        desc = r[idx["Description"]]

        is_target = target_matches_fields(
            target,
            [protein, gene, desc],
            allow_substring=True,
        )

        if is_target:
            rows.append({c: r[idx[c]] if c in idx else None for c in wanted})

    wb.close()

    out = pd.DataFrame(rows)
    if out.empty:
        print(f"WARNING: no rows matching {target.query} found in semitryptome.")
        return out

    out["dataset"] = "Semitryptome"
    out["cleavage_site"] = safe_numeric(out["CleavageSite"]).astype("Int64")
    out["psm"] = safe_numeric(out["PSM"])
    out["samples"] = safe_numeric(out["#Samples"])
    out["peptide"] = out["Peps"].astype(str).str.split(";").str[0].map(clean_peptide_sequence)
    out["start"], out["end"] = zip(*[
        peptide_span_from_start(p, s) for p, s in zip(out["peptide"], out["cleavage_site"])
    ])
    out["weight"] = out["psm"].fillna(1)

    return out

# out["weight"] = out["max_abs_logfc"].fillna(0) for plotting
# larger max_abs_logFC = larger HUNTER dot
# out["max_abs_logfc"] = out[logfc_present].abs().max(axis=1, skipna=True)
# We use this strategy because the HUNTER dataset has several methyl-jasmonate time points,
# and we need one simple number to represent how strongly each peptide responds.
# For each peptide row, look across all MJ logFC time points, ignore the sign, and take the strongest change.
# abs: This is useful because both strong increase and strong decrease indicate a strong treatment-associated change.
# HUNTER dot position = peptide start / cleavage coordinate
# HUNTER dot size = strongest MJ response at any time point

# Explanation: Because HUNTER contains several methyl-jasmonate time points, 
# I summarized each peptide’s response using the maximum absolute log2 fold-change across time. 
# This gives one response-strength value per peptide and allows marker size in the plot to represent the strongest treatment-associated change. 
# I used the absolute value because both strong increases and strong decreases indicate responsive N-terminal peptides.
def read_hunter_arabidopsis(path: Path, target: TargetSpec, sheet_name: str = "Only significant peptides") -> pd.DataFrame:
    """
    Read Arabidopsis HUNTER methyl-jasmonate data and retain rows matching the target protein.

    If the requested sheet is absent, the script tries 'All peptides'.
    """
    wb = load_workbook(path, read_only=True, data_only=True)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "All peptides" in wb.sheetnames:
        print(f"WARNING: sheet '{sheet_name}' not found in {path.name}; using 'All peptides'.")
        ws = wb["All peptides"]
    else:
        wb.close()
        raise ValueError(f"Could not find HUNTER sheet in {path.name}. Available sheets: {wb.sheetnames}")

    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = unique_headers(raw_headers)
    idx = {h: i for i, h in enumerate(headers)}

    required = ["protein", "gene", "peptide"]
    missing = [c for c in required if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Missing HUNTER columns in {path.name}: {missing}")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        protein = r[idx["protein"]]
        gene = r[idx["gene"]]

        is_target = target_matches_fields(
            target,
            [protein, gene],
            allow_substring=True,
        )

        if is_target:
            rows.append({h: r[i] if i < len(r) else None for i, h in enumerate(headers)})

    wb.close()

    out = pd.DataFrame(rows)
    if out.empty:
        print(f"WARNING: no rows matching {target.query} found in HUNTER data.")
        return out

    out["dataset"] = "HUNTER MJ"

    # In some HUNTER tables, "cleavage_site" contains a sequence string and
    # numeric coordinates are stored in "first_aa_position".
    cs = safe_numeric(out["cleavage_site"]) if "cleavage_site" in out.columns else pd.Series(np.nan, index=out.index)
    if cs.isna().all() and "first_aa_position" in out.columns:
        cs = safe_numeric(out["first_aa_position"])

    out["cleavage_site"] = cs.astype("Int64")
    out["peptide"] = out["peptide"].astype(str).map(clean_peptide_sequence)
    out["start"], out["end"] = zip(*[
        peptide_span_from_start(p, s) for p, s in zip(out["peptide"], out["cleavage_site"])
    ])

    for c in LOGFC_COLS:
        if c in out.columns:
            out[c] = safe_numeric(out[c])

    logfc_present = [c for c in LOGFC_COLS if c in out.columns]
    if logfc_present:
        out["max_abs_logfc"] = out[logfc_present].abs().max(axis=1, skipna=True)
        out["mean_logfc"] = out[logfc_present].mean(axis=1, skipna=True)
    else:
        out["max_abs_logfc"] = np.nan
        out["mean_logfc"] = np.nan

    out["weight"] = out["max_abs_logfc"].fillna(0)

    return out

# calculate one maximum metacaspase score per row : Target protein AT3G09260.1
# Start = 284
# End = 293
# Sequence = DSQDGASIDR
# out["metacaspase_score_max"] = out[score_cols].max(axis=1, skipna=True)
# Because the Data sheet contains multiple metacaspase-related score sections. 
# A peptide may be supported in one section but not another.
# For plotting, we need one value per row. The maximum score answers:
# What is the strongest metacaspase evidence for this peptide window?
# larger metacaspase_score_max = larger metacaspase dot
# larger dot = this peptide/start site has stronger metacaspase evidence in at least one score column
# The task is asking for a visualization that makes overlap/proximity intuitive. It is not asking for a detailed comparison of each individual metacaspase.
# So for the main plot, using the maximum score is fine because it reduces many score columns to one interpretable value:
# So when we say “different MC,” we mean different Arabidopsis metacaspase enzymes or metacaspase assay groups.
# metacaspase = enzyme that can cleave proteins

# Explain: The metacaspase matrix contains several score columns, likely corresponding to different metacaspase enzymes or experimental score blocks. 
# For the integrated overview plot, I summarized them by taking the maximum score per peptide window. This gives one metacaspase evidence value per site. 
# It should be interpreted as evidence from at least one metacaspase-related score, not as evidence that all metacaspases cleave the site.
def read_metacaspase(path: Path, target: TargetSpec, match_isoforms: bool = False) -> pd.DataFrame:
    """
    Read the ProteinBased_www_P10'P10 metacaspase matrix and retain only rows
    whose PRIMARY Protein column is the requested target protein ID.

    Strict rule:
      keep a row only if Protein == target.query after light normalization.

    This intentionally does NOT use aliases, UniProt ID, Description, or Isoforms
    for metacaspase extraction. Those fields can contain shared or homologous
    mappings, which would incorrectly pull rows from related BGLU proteins such
    as AT1G66270.2, AT1G66280.1, AT3G16420.2, or AT3G21370.1 when the requested
    target is AT3G09260.1.

    The match_isoforms argument is retained only for backward compatibility with
    older GUI code; it is ignored in this strict extraction function.
    """
    if match_isoforms:
        print(
            "WARNING: match_isoforms=True was requested, but metacaspase extraction "
            "is now strict and uses only the primary Protein column."
        )

    wb = load_workbook(path, read_only=True, data_only=True)

    if "Data" in wb.sheetnames:
        ws = wb["Data"]
    else:
        ws = wb[wb.sheetnames[0]]
        print(f"WARNING: sheet 'Data' not found in {path.name}; using '{ws.title}'.")

    header_row = 4
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    headers = unique_headers(raw_headers)
    idx = {h: i for i, h in enumerate(headers)}

    required = ["Protein", "Description", "Start", "End", "Sequence"]
    missing = [c for c in required if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Missing metacaspase columns in {path.name}: {missing}")

    target_primary_id = normalize_match_text(target.query)
    if not target_primary_id:
        wb.close()
        raise ValueError("The primary target ID is empty; metacaspase extraction requires a Protein-column ID.")

    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        primary_protein = r[idx["Protein"]]
        primary_protein_norm = normalize_match_text(primary_protein)

        # Strict primary-column match only. Do not search Isoforms, Description,
        # UniProt accession, or gene aliases here.
        if primary_protein_norm == target_primary_id:
            rows.append({h: r[i] if i < len(r) else None for i, h in enumerate(headers)})

    wb.close()

    out = pd.DataFrame(rows)
    if out.empty:
        print(f"WARNING: no primary Protein == {target.query} rows found in metacaspase matrix.")
        return out

    out["dataset"] = "Metacaspase matrix"
    out["cleavage_site"] = safe_numeric(out["Start"]).astype("Int64")
    out["peptide"] = out["Sequence"].astype(str).map(clean_peptide_sequence)
    out["start"] = safe_numeric(out["Start"])
    out["end"] = safe_numeric(out["End"])

    score_cols = [c for c in out.columns if str(c).startswith("Score")]
    if score_cols:
        for c in score_cols:
            out[c] = safe_numeric(out[c])
        out["metacaspase_score_max"] = out[score_cols].max(axis=1, skipna=True)
    else:
        out["metacaspase_score_max"] = np.nan

    out["weight"] = out["metacaspase_score_max"].fillna(0)

    print("  Metacaspase final extracted Protein values:")
    for protein_id in sorted(out["Protein"].astype(str).unique()):
        print(f"    - {protein_id}")

    return out

# ---------------------------------------------------------------------
# Lollipop plotting
# ---------------------------------------------------------------------

# | Dataset            | Weight means              |
# | ------------------ | ------------------------- |
# | Semitryptome       | PSM                       |
# | HUNTER MJ          | maximum absolute logFC    |
# | Metacaspase matrix | maximum metacaspase score |

# So if the same dataset has multiple rows at the same coordinate, 
# the plot will show one dot, not many overlapping dots.

# Because PSM gives peptide-spectrum match support. A higher PSM means stronger peptide evidence.
# for example: 
# | cleavage_site | dataset            | weight |
# | ------------: | ------------------ | -----: |
# |            23 | Semitryptome       |      5 |
# |            25 | HUNTER MJ          |    2.3 |
# |           284 | Metacaspase matrix |   0.61 |

# So rows are collapsed only if they come from the same dataset and have the same exact cleavage coordinate.
# Because the same coordinate may appear in multiple datasets, and we want to keep those separate.
def build_lollipop_points(
    semi: pd.DataFrame,
    hunter: pd.DataFrame,
    meta: pd.DataFrame,
    aggregate_sites: bool = True,
) -> pd.DataFrame:
    """
    Convert the three extracted PYK10 tables into one plotting table.

    Weight meaning:
      Semitryptome: PSM
      HUNTER MJ: maximum absolute logFC
      Metacaspase matrix: maximum metacaspase score

    By default this function collapses exact duplicate coordinates, so the plot
    has one dot per dataset per cleavage coordinate. This avoids a misleading
    plot in which repeated peptide observations at the same coordinate look like
    many independent cleavage sites.
    """
    parts = []

    if not semi.empty:
        tmp = semi.copy()
        tmp["dataset"] = "Semitryptome"
        tmp["weight"] = safe_numeric(tmp.get("psm", tmp.get("PSM", 1))).fillna(1)
        parts.append(tmp[["cleavage_site", "dataset", "weight", "peptide", "start", "end"]])

    if not hunter.empty:
        tmp = hunter.copy()
        tmp["dataset"] = "HUNTER MJ"
        tmp["weight"] = safe_numeric(tmp.get("max_abs_logfc", 0)).fillna(0)
        parts.append(tmp[["cleavage_site", "dataset", "weight", "peptide", "start", "end"]])

    if not meta.empty:
        tmp = meta.copy()
        tmp["dataset"] = "Metacaspase matrix"
        tmp["weight"] = safe_numeric(tmp.get("metacaspase_score_max", 0)).fillna(0)
        parts.append(tmp[["cleavage_site", "dataset", "weight", "peptide", "start", "end"]])

    if not parts:
        return pd.DataFrame(columns=["cleavage_site", "dataset", "weight", "peptide", "start", "end"])

    combined = pd.concat(parts, ignore_index=True)
    combined["cleavage_site"] = safe_numeric(combined["cleavage_site"])
    combined = combined.dropna(subset=["cleavage_site"]).copy()
    combined["cleavage_site"] = combined["cleavage_site"].astype(int)
    combined["weight"] = safe_numeric(combined["weight"]).fillna(0)

    if not aggregate_sites:
        combined["n_observations"] = 1
        return combined

    def first_nonempty(x):
        vals = [str(v) for v in x if str(v) and str(v).lower() != "nan"]
        return ";".join(vals[:3]) if vals else ""

    aggregated = (
        combined
        .groupby(["dataset", "cleavage_site"], as_index=False)
        .agg(
            weight=("weight", "max"),
            n_observations=("weight", "size"),
            peptide=("peptide", first_nonempty),
            start=("start", "min"),
            end=("end", "max"),
        )
    )
    return aggregated


def build_cluster_summary(points: pd.DataFrame, gap: int = 10) -> pd.DataFrame:
    """
    Group nearby cleavage sites into processing hotspots.

    Rule:
      After sorting unique cleavage positions, a position joins the current cluster
      if it is within `gap` amino acids of the previous site.
    """
    if points.empty:
        return pd.DataFrame()

    sites = sorted(points["cleavage_site"].unique())

    clusters = []
    current = []
    for site in sites:
        if not current or site - current[-1] <= gap:
            current.append(site)
        else:
            clusters.append(current)
            current = [site]
    if current:
        clusters.append(current)

    rows = []
    for i, cluster_sites in enumerate(clusters, start=1):
        start = min(cluster_sites)
        end = max(cluster_sites)
        sub = points[points["cleavage_site"].between(start, end)]

        support = {}
        supporting_datasets = 0
        for dataset in ["Semitryptome", "HUNTER MJ", "Metacaspase matrix"]:
            n = sub[sub["dataset"] == dataset]["cleavage_site"].nunique()
            support[dataset] = int(n)
            if n > 0:
                supporting_datasets += 1

        rows.append({
            "cluster_id": f"C{i}",
            "cluster_start": start,
            "cluster_end": end,
            "cluster_center": float(np.mean(cluster_sites)),
            "sites": ";".join(map(str, cluster_sites)),
            "datasets_supporting": supporting_datasets,
            "Semitryptome": support["Semitryptome"],
            "HUNTER MJ": support["HUNTER MJ"],
            "Metacaspase matrix": support["Metacaspase matrix"],
            "n_total_sites": len(cluster_sites),
        })

    return pd.DataFrame(rows)

def infer_protein_length(semi: pd.DataFrame, hunter: pd.DataFrame, meta: pd.DataFrame) -> int:
    """
    Infer plotting length from peptide ends and HUNTER relative positions.
    """
    length_candidates = []

    if not hunter.empty and "preceding_aa_position_relative" in hunter.columns and "first_aa_position" in hunter.columns:
        rel = safe_numeric(hunter["preceding_aa_position_relative"])
        first = safe_numeric(hunter["first_aa_position"])
        valid = (rel > 0) & first.notna()
        length_candidates.extend((first[valid] / (rel[valid] / 100)).dropna().tolist())

    max_end = 0
    for df in [semi, hunter, meta]:
        if not df.empty and "end" in df.columns:
            m = safe_numeric(df["end"]).max()
            if pd.notna(m):
                max_end = max(max_end, int(m))

    if length_candidates:
        estimated = int(round(float(np.nanmedian(length_candidates))))
        return max(estimated, max_end, DEFAULT_LENGTH)

    return max(max_end, DEFAULT_LENGTH)

# instead of treating every nearby site as a completely independent biological event.
# The reason is that cleavage evidence often appears as several nearby starts, not one perfect exact residue. 
# So this function helps us summarize the region-level evidence while still showing the exact coordinates in the tables.
# a new site joins the same cluster if it is within 10 amino acids of the previous site. 
# because each new site is close to the previous one. 
# This creates clusters of nearby sites that likely represent the same underlying processing event or region, rather than treating them as completely independent events.
# Clusters are used as a visual summary of proximal cleavage evidence, not as proof of a single biochemical cleavage site.
def plot_lollipop_multisource_hotspots(
    points: pd.DataFrame,
    cluster_summary: pd.DataFrame,
    protein_len: int,
    out_png: Path,
    out_pdf: Path,
    shade_min_dataset_support: int = 2,
    x_tick_gap: int = 20,
    title: str = "Target protein: cleavage-site evidence",
    x_label: str = "Residue coordinate",
) -> None:
    """
    Generate a clean presentation-ready lollipop plot.

    Features:
      - one horizontal track per dataset
      - dot at each cleavage coordinate
      - marker size scaled by evidence strength within each dataset
      - faint lollipop stems
      - background shading for clusters/hotspots supported by >= 2 datasets
      - configurable x-axis tick spacing using x_tick_gap
    """
    if points.empty:
        raise ValueError("No points available for lollipop plot.")

    # Compact vertical spacing
    y_positions = {
        "Semitryptome": 1.45,
        "HUNTER MJ": 1.10,
        "Metacaspase matrix": 0.80
    }

    dataset_order = ["Semitryptome", "HUNTER MJ", "Metacaspase matrix"]
    cluster_label_y = max(y_positions.values()) + 0.15

    fig, ax = plt.subplots(figsize=(15, 5.8))

    # Background cluster shading
    if not cluster_summary.empty:
        shaded = cluster_summary[
            cluster_summary["datasets_supporting"] >= shade_min_dataset_support
        ]

        for _, row in shaded.iterrows():
            ax.axvspan(
                row["cluster_start"],
                row["cluster_end"],
                alpha=0.20,
                zorder=0,
            )

            ax.text(
                row["cluster_center"],
                cluster_label_y,
                row["cluster_id"],
                ha="center",
                va="bottom",
                fontsize=8,
            )

    # Lollipop tracks
    for dataset in dataset_order:
        y = y_positions[dataset]
        sub = points[points["dataset"] == dataset].copy()

        if sub.empty:
            continue

        max_weight = sub["weight"].max()
        if pd.isna(max_weight) or max_weight <= 0:
            max_weight = 1.0

        # Scale marker size within each dataset
        sub["marker_size"] = 30 + 85 * (sub["weight"] / max_weight)

        # Faint lollipop stems
        for x in sub["cleavage_site"]:
            ax.plot(
                [x, x],
                [y - 0.12, y + 0.12],
                linewidth=0.65,
                alpha=0.30,
                zorder=1,
            )

        # Lollipop dots
        ax.scatter(
            sub["cleavage_site"],
            np.full(len(sub), y),
            s=sub["marker_size"],
            alpha=0.75,
            label=dataset,
            zorder=2,
        )

    # X-axis: protein coordinate
    ax.set_xlim(0, protein_len + 10)
    ax.set_xticks(np.arange(0, protein_len + x_tick_gap, x_tick_gap))
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right")

    # Y-axis: dataset tracks
    ax.set_ylim(0.65, cluster_label_y + 0.15)
    ax.set_yticks([y_positions[d] for d in dataset_order])
    ax.set_yticklabels(dataset_order)

    ax.set_xlabel(x_label)
    ax.set_title(title, pad=5)

    ax.grid(axis="x", alpha=0.23)

    # Legend above the plot
    ax.legend(
        loc="upper center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.76, 1.0),
        fontsize=12,)

    fig.tight_layout()
    fig.savefig(out_png, dpi=600)
    fig.savefig(out_pdf)
    plt.close(fig)

#  python scripts/run_analysis_lollipop_final.py --input-dir data --output-dir plot_aggregated_meta

def plot_cluster_matrix(cluster_summary: pd.DataFrame, out_png: Path, out_pdf: Path, title: str = "Cluster-level evidence matrix") -> None:
    """
    Optional companion plot: simple matrix showing which clusters are supported by which datasets.
    """
    if cluster_summary.empty:
        return

    keep = cluster_summary[
        (cluster_summary["datasets_supporting"] >= 2)
        | (cluster_summary["HUNTER MJ"] > 0)
    ].copy()

    if keep.empty:
        keep = cluster_summary.copy()

    dataset_cols = ["Semitryptome", "HUNTER MJ", "Metacaspase matrix"]
    keep["label"] = keep.apply(
        lambda r: f"{r['cluster_id']} ({int(r['cluster_start'])}-{int(r['cluster_end'])})",
        axis=1,
    )

    fig, ax = plt.subplots(figsize=(8.5, max(4.5, 0.35 * len(keep))))

    for i, (_, row) in enumerate(keep.iterrows()):
        for j, dataset in enumerate(dataset_cols):
            value = int(row[dataset])
            size = 25 + 70 * value if value > 0 else 8
            ax.scatter(j, i, s=size, alpha=0.75 if value else 0.20)
            if value:
                ax.text(j, i, str(value), ha="center", va="center", fontsize=8)

    ax.set_xticks(range(len(dataset_cols)))
    ax.set_xticklabels(dataset_cols)
    ax.set_yticks(range(len(keep)))
    ax.set_yticklabels(keep["label"])
    ax.invert_yaxis()
    ax.set_xlabel("Dataset")
    ax.set_ylabel("Cleavage-region cluster")
    ax.set_title(title)
    ax.grid(alpha=0.2)
    fig.savefig(out_png, dpi=300)
    fig.savefig(out_pdf)
    plt.close(fig)

def write_readme(
    output_dir: Path,
    target: TargetSpec,
    sem: pd.DataFrame,
    hunter: pd.DataFrame,
    meta: pd.DataFrame,
    points: pd.DataFrame,
    cluster_summary: pd.DataFrame,
    cluster_gap: int,
) -> None:
    alias_text = ", ".join(target.terms)
    md = f"""# {target.display_name} lollipop cleavage-site visualization

This output was generated by `run_analysis_lollipop.py`.

## Objective

Visualize cleavage-site evidence for **{target.display_name}** as a clean lollipop plot, integrating three Arabidopsis datasets:

1. Arabidopsis semitryptome
2. HUNTER shoot methyl-jasmonate N-terminomics
3. Protein-based metacaspase substrate matrix

## Target matching

The target was searched using the following identifiers/names:

`{alias_text}`

Rows were retained when one of these terms matched the relevant identifier or descriptive fields in each dataset. For the metacaspase matrix, matching was restricted to the primary `Protein` column unless `--metacaspase-include-isoforms` was used.

## Extracted rows

| Dataset | Rows extracted |
|---|---:|
| Semitryptome | {len(sem)} |
| HUNTER MJ | {len(hunter)} |
| Metacaspase matrix | {len(meta)} |
| Combined lollipop points | {len(points)} |
| Cleavage-site clusters | {len(cluster_summary)} |

## Plot interpretation

- **X-axis**: residue coordinate in the selected target protein.
- **Y-axis**: dataset source.
- **Dot position**: exact cleavage or peptide-start coordinate.
- **Dot size**: evidence strength.
  - Semitryptome: PSM support.
  - HUNTER MJ: maximum absolute log2 fold-change.
  - Metacaspase matrix: maximum substrate score.
- **Shaded bands**: cleavage-site clusters supported by at least two datasets.

Clusters were defined using a simple reproducible rule: adjacent cleavage sites are assigned to the same cluster when they are within **{cluster_gap} amino acids** of the previous site after sorting.

## Main files

- `plots/{target.output_prefix}_lollipop_multisource_hotspots.png`
- `plots/{target.output_prefix}_lollipop_multisource_hotspots.pdf`
- `plots/{target.output_prefix}_cluster_evidence_matrix.png`
- `tables/{target.output_prefix}_lollipop_points.csv`
- `tables/{target.output_prefix}_lollipop_clusters.csv`

## Recommended presentation wording

Exact cleavage coordinates were retained, but nearby sites were additionally grouped into reproducible {cluster_gap}-aa clusters to highlight processing regions. This avoids overinterpreting adjacent N-terminal peptide starts as independent events while preserving the residue-level evidence in the output tables.
"""
    (output_dir / "README_lollipop.md").write_text(md, encoding="utf-8")


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a lollipop cleavage-site evidence plot for any target Arabidopsis protein."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("data"),
        help="Directory containing the input Excel files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output_lollipop"),
        help="Directory for output tables and plots.",
    )
    parser.add_argument(
        "--target",
        default=DEFAULT_TARGET,
        help=(
            "Primary target protein/gene/name to search. For Arabidopsis this is usually "
            "the TAIR protein ID, for example AT3G09260.1. Default keeps the original PYK10 analysis."
        ),
    )
    parser.add_argument(
        "--uniprot-id",
        default=None,
        help=(
            "Optional UniProt accession for the same protein, provided separately because "
            "some source files use UniProt IDs while others use TAIR IDs. Example: Q9SR37."
        ),
    )
    parser.add_argument(
        "--aliases",
        nargs="*",
        default=[],
        help=(
            "Additional gene/protein names for the same target. Separate values with spaces, "
            "or use comma/semicolon-separated strings. Example: --aliases PYK10 BGLU23 LEB"
        ),
    )
    parser.add_argument(
        "--alias-map",
        type=Path,
        default=None,
        help=(
            "Optional CSV/TSV mapping file with columns such as target, aliases, length, "
            "display_name, output_prefix. This is useful when one protein has several IDs."
        ),
    )
    parser.add_argument(
        "--protein-length",
        type=int,
        default=None,
        help="Optional known protein length. If omitted, the script infers a plotting length when possible.",
    )
    parser.add_argument(
        "--output-prefix",
        default=None,
        help="Optional filename prefix. By default, this is derived from --target.",
    )
    parser.add_argument(
        "--metacaspase-include-isoforms",
        action="store_true",
        help=(
            "Deprecated/ignored. Metacaspase extraction is always strict: primary Protein column == --target only."
        ),
    )
    parser.add_argument(
        "--cluster-gap",
        type=int,
        default=10,
        help="Maximum gap in amino acids for grouping adjacent sites into clusters.",
    )
    parser.add_argument(
        "--shade-min-dataset-support",
        type=int,
        default=2,
        help="Minimum number of datasets supporting a cluster for background shading.",
    )
    parser.add_argument(
        "--plot-every-row",
        action="store_true",
        help=(
            "Plot every extracted row instead of aggregating exact duplicate cleavage "
            "positions. Default is one dot per dataset per cleavage coordinate."
        ),
    )
    args = parser.parse_args()

    target = make_target_spec(
        query=args.target,
        uniprot_id=args.uniprot_id,
        aliases=args.aliases,
        alias_map=args.alias_map,
        protein_length=args.protein_length,
        output_prefix=args.output_prefix,
    )

    plots_dir, tables_dir = ensure_output_dirs(args.output_dir)
    paths = resolve_input_files(args.input_dir)

    print(f"\nTarget search terms: {', '.join(target.terms)}")
    print(f"Display name: {target.display_name}")
    print(f"Output prefix: {target.output_prefix}")

    print(f"\nReading and extracting rows for {target.query}...")
    sem = read_semitryptome(paths["semitryptome"], target)
    hunter = read_hunter_arabidopsis(paths["hunter_arabidopsis"], target)
    meta = read_metacaspase(
        paths["metacaspase"],
        target,
        match_isoforms=args.metacaspase_include_isoforms,
    )

    print(f"  Semitryptome rows: {len(sem)}")
    print(f"  HUNTER MJ rows: {len(hunter)}")
    print(f"  Metacaspase matrix rows: {len(meta)}")

    def _n_unique_sites(df):
        if df.empty or "cleavage_site" not in df.columns:
            return 0
        return safe_numeric(df["cleavage_site"]).dropna().astype(int).nunique()

    print("  Unique exact cleavage coordinates:")
    print(f"    Semitryptome: {_n_unique_sites(sem)}")
    print(f"    HUNTER MJ: {_n_unique_sites(hunter)}")
    print(f"    Metacaspase matrix: {_n_unique_sites(meta)}")

    prefix = target.output_prefix

    # Save extracted source tables.
    sem.to_csv(tables_dir / f"{prefix}_semitryptome.csv", index=False)
    hunter.to_csv(tables_dir / f"{prefix}_hunter_mj.csv", index=False)
    meta.to_csv(tables_dir / f"{prefix}_metacaspase.csv", index=False)

    points = build_lollipop_points(sem, hunter, meta, aggregate_sites=not args.plot_every_row)
    cluster_summary = build_cluster_summary(points, gap=args.cluster_gap)
    protein_len = target.protein_length or infer_protein_length(sem, hunter, meta)

    points.to_csv(tables_dir / f"{prefix}_lollipop_points.csv", index=False)
    cluster_summary.to_csv(tables_dir / f"{prefix}_lollipop_clusters.csv", index=False)

    print(f"  Lollipop points: {len(points)}")
    print(f"  Clusters: {len(cluster_summary)}")
    print(f"  Protein plotting length: {protein_len}")

    plot_title = f"{target.display_name}: cleavage-site evidence"
    x_label = f"{target.query} residue coordinate"

    print("\nGenerating plots...")
    main_png = plots_dir / f"{prefix}_lollipop_multisource_hotspots.png"
    main_pdf = plots_dir / f"{prefix}_lollipop_multisource_hotspots.pdf"
    matrix_png = plots_dir / f"{prefix}_cluster_evidence_matrix.png"
    matrix_pdf = plots_dir / f"{prefix}_cluster_evidence_matrix.pdf"

    plot_lollipop_multisource_hotspots(
        points=points,
        cluster_summary=cluster_summary,
        protein_len=protein_len,
        out_png=main_png,
        out_pdf=main_pdf,
        shade_min_dataset_support=args.shade_min_dataset_support,
        title=plot_title,
        x_label=x_label,
    )

    plot_cluster_matrix(
        cluster_summary=cluster_summary,
        out_png=matrix_png,
        out_pdf=matrix_pdf,
        title=f"{target.query} cluster-level evidence matrix",
    )

    write_readme(
        output_dir=args.output_dir,
        target=target,
        sem=sem,
        hunter=hunter,
        meta=meta,
        points=points,
        cluster_summary=cluster_summary,
        cluster_gap=args.cluster_gap,
    )

    print("\nDone.")
    print(f"Output directory: {args.output_dir}")
    print(f"Main plot: {main_png}")
    print(f"Cluster matrix: {matrix_png}")


if __name__ == "__main__":
    main()
